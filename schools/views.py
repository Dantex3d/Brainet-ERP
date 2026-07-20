import datetime
from datetime import timedelta
import email
import re
from urllib.parse import urlparse
from django.utils import timezone
from django.conf import settings
import logging
from io import BytesIO
import openpyxl
from utils.email_service import send_email

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, get_user_model, login
from django.db import transaction
from exams.models import Exam, Mark
from django.db.models import Q
from django.urls import reverse

import students
import subjects
from .models import DOSMessage, DOSQuery, Notification, School, DirectorOfStudies, Dormitory, Term, Class, Subject, GradingPolicy, StudentMark, StudentPromotion, SchoolNotice, ErrorReport, SecurityLog, ContactMessage
from django.db import IntegrityError
from collections import defaultdict
from students.models import Student
from schools.models import School, Dormitory, DirectorOfStudies, Term
from schools.models import Class, Subject, VoucherRequest
from classes.models import Stream
from teachers.models import ClassTeacherAssignment, Teacher, TeacherSubjectAssignment, OnlineClass, OnlineClassParticipant
from subjects.models import ClassSubject
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.platypus import Image as RLImage
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.shapes import Drawing, Circle

User = get_user_model()


def get_superuser_emails():
    return list(
        User.objects.filter(is_superuser=True, email__isnull=False)
        .values_list('email', flat=True)
    )


def resolve_exam_window_state(request, school):
    """Return the current exam-window state, preferring the latest session action when present."""
    if not school:
        return False, 0

    active_exams = Exam.objects.filter(school=school, is_active=True)
    session_key = f"exam_window_state_{getattr(school, 'id', None)}"

    if session_key in request.session:
        is_open = request.session[session_key] == "open"
        return is_open, active_exams.count()

    return active_exams.exists(), active_exams.count()


def set_exam_window_state(request, school, is_open):
    """Keep the latest exam-window action in the session for compatibility with other views."""
    if not school:
        return

    session_key = f"exam_window_state_{school.id}"
    request.session[session_key] = "open" if is_open else "closed"
    request.session.modified = True


def clear_exam_window_state(request, school):
    """Remove any stale exam-window session override for the current school."""
    if not school:
        return

    session_key = f"exam_window_state_{school.id}"
    request.session.pop(session_key, None)
    request.session.modified = True


def get_school_login_domain(school):
    """Return a short sanitized login domain for a school."""
    if not school or not getattr(school, "name", None):
        return "school"

    words = re.findall(r"[a-z0-9]+", school.name.lower())
    if not words:
        return "school"

    first = words[0][:5]
    last = words[-1]

    generic_suffixes = {
        "school",
        "academy",
        "college",
        "institute",
        "primary",
        "infants",
        "nursery",
        "high",
        "junior",
    }

    if len(words) > 1 and last in generic_suffixes:
        second = words[-2][:3]
        return f"{first}{second}"

    second = last[:3]
    return f"{first}{second}"


def student_login_email(admission_number, school):
    return f"{admission_number}@{get_school_login_domain(school)}.school"


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from users.models import CustomUser
from teachers.models import Teacher


def superuser_required(view_func):
    return user_passes_test(lambda u: u.is_superuser)(view_func)


def exam_controller_required(view_func):
    return user_passes_test(
        lambda u: u.is_superuser or getattr(u, "role", None) in ["dos", "principal"]
    )(view_func)

from django.conf import settings
from django.urls import reverse
from utils.email_service import send_email
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.core.files.storage import default_storage

def format_position_display(value):
    """Render ranking positions without dash placeholders."""
    if value is None:
        return ""

    if isinstance(value, str):
        value = value.strip()
        if not value or value in {"-", "—", "None", "none"}:
            return ""
        return value

    if isinstance(value, (int, float)):
        return str(int(value)) if float(value).is_integer() else str(value)

    return str(value)


def assign_competition_ranks(items, score_getter, rank_attr="rank"):
    """Assign competition-style ranks where ties share the same position and the next rank skips."""
    ranked_items = sorted(items, key=lambda item: score_getter(item), reverse=True)

    prev_score = None
    prev_rank = None
    for index, item in enumerate(ranked_items, start=1):
        score = score_getter(item)
        if prev_score is not None and score == prev_score:
            rank = prev_rank
        else:
            rank = index
            prev_rank = rank
        prev_score = score

        if isinstance(item, dict):
            item[rank_attr] = rank
        else:
            setattr(item, rank_attr, rank)

    return ranked_items


logger = logging.getLogger(__name__)





def send_user_verification_email(user, request=None, role_name=None):
    # Ensure user has an email
    if not getattr(user, 'email', None):
        return False

    # Resolve display name: prefer full name, then first_name, then related profile names
    display_name = None
    try:
        display_name = user.get_full_name() if user.get_full_name() else None
    except Exception:
        display_name = getattr(user, 'first_name', None)

    if not display_name:
        # Try related profiles (teacher/principal/dos)
        try:
            from teachers.models import Teacher
            t = Teacher.objects.filter(user=user).first()
            if t and getattr(t, 'name', None):
                display_name = t.name
        except Exception:
            pass

    if not display_name:
        display_name = user.email

    # =========================
    # DOS / PRINCIPAL FLOW (6-digit code)
    # =========================
    if user.role in ['dos', 'principal']:
        code = user.generate_verification_code()
        verify_link = (
            request.build_absolute_uri(reverse('verify_user_code')) + f"?email={user.email}"
            if request
            else reverse('verify_user_code') + f"?email={user.email}"
        )

        subject = f"Verify your {role_name or user.role} account on Brainet"

        html_body = f"<p>Dear {display_name},</p>"
        html_body += f"<p>Your {role_name or user.role} account has been created on Brainet.</p>"
        html_body += f"<p>Please use the verification code below (expires in 1 hour):</p>"
        html_body += f"<h2>{code}</h2>"
        html_body += f"<p>Or click this link to open the verification page:</p>"
        html_body += f"<p><a href=\"{verify_link}\">Open Verification Page</a></p>"
        html_body += f"<p><strong>Important:</strong> This link only opens the verification page; you still must enter the 6-digit code above.</p>"
        html_body += f"<p>If you did not request this, contact support.</p>"

        sent = send_email(to_email=user.email, subject=subject, message=html_body, recipient_name=display_name, html=True)

        superuser_emails = get_superuser_emails()
        if superuser_emails:
            su_subject = f"New {role_name or user.role.title()} verification sent for {display_name}"
            su_body = f"<p>Dear Superuser,</p>"
            su_body += f"<p>A new {role_name or user.role} account was created for {display_name} ({user.email}).</p>"
            su_body += f"<p>Verification code: <strong>{code}</strong></p>"
            su_body += f"<p>Verification page: <a href=\"{verify_link}\">{verify_link}</a></p>"
            su_body += f"<p>The link opens the verification page only; the 6-digit code must still be entered to complete verification.</p>"
            send_email(to_email=superuser_emails, subject=su_subject, message=su_body, recipient_name='Superuser', html=True)

        return sent

    # =========================
    # OTHER USERS FLOW (both code + email token link)
    # =========================
    # Generate both a short code and a token link so users may verify either way
    token = user.generate_email_verification_token()
    code = user.generate_verification_code()

    verify_link = (
        request.build_absolute_uri(reverse('verify_user_email', args=[token]))
        if request
        else reverse('verify_user_email', args=[token])
    )

    subject = f"Verify your {role_name or 'Brainet'} account"

    html_body = f"<p>Hi {display_name},</p>"
    html_body += f"<p>We've created your account on Brainet. You may verify your email using either the verification code below, or by clicking the verification link.</p>"
    html_body += f"<p><strong>Verification code:</strong></p>"
    html_body += f"<h2 style=\"letter-spacing:4px;\">{code}</h2>"
    html_body += f"<p>Or click the button to verify now:</p>"
    html_body += f"<p><a href=\"{verify_link}\" style=\"display:inline-block;padding:10px 18px;background:#0d6efd;color:#fff;border-radius:6px;text-decoration:none;\">Verify Email</a></p>"
    html_body += f"<p>If you didn't request this, ignore this email.</p>"

    return send_email(to_email=user.email, subject=subject, message=html_body, recipient_name=display_name, html=True)

def send_school_verification_email(school, request=None):
    if not school.email:
        return

    token = school.generate_verification_token()
    verify_link = request.build_absolute_uri(
        reverse('verify_school_via_token', args=[token])
    ) if request else reverse('verify_school_via_token', args=[token])

    subject = "Verify your school registration on Brainet"
    body = (
        f"Hello {school.name},\n\n"
        f"Please verify this school's email address by clicking the link below:\n\n{verify_link}\n\n"
        "This link expires in 1 hour. After verification the school admin will be able to complete activation and password resets.\n\n"
        "If you did not register this school, please ignore this message."
    )

    return send_email(
        to_email=school.email,
        subject=subject,
        message=body,
        recipient_name=school.name,
        html=False,
    )


@superuser_required
def pending_verification(request):
    """Superuser view: list users pending email verification with search and filters."""
    q = request.GET.get('q', '').strip()
    filter_type = request.GET.get('filter', '')

    users = User.objects.filter(email_verified=False)

    if q:
        users = users.filter(Q(email__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q))

    if filter_type == 'failed':
        users = users.filter(verification_attempts__gt=0)

    users = users.order_by('-email_verification_sent_at')

    paginator = Paginator(users, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'schools/pending_verification.html', {
        'users': page_obj,
        'q': q,
        'filter': filter_type,
    })


@superuser_required
def pending_verification_count(request):
    count = User.objects.filter(email_verified=False).count()
    return JsonResponse({'count': count})


@superuser_required
def resend_verification_email(request, user_id):
    user = get_object_or_404(User, id=user_id)
    # Use send_user_verification_email's boolean result to determine success/failure
    try:
        sent = send_user_verification_email(user, request=request)
        if sent:
            messages.success(request, f"Verification email resent to {user.email}.")
        else:
            support = getattr(settings, 'SUPPORT_EMAIL', None) or getattr(settings, 'DEFAULT_FROM_EMAIL', None) or 'support@brainet.local'
            messages.error(request, f"Could not resend verification to {user.email}. Contact: {support}")
    except Exception as e:
        messages.error(request, f"Could not resend verification: {str(e)}")
    return redirect('pending_verification')


def contact_submit(request):
    if request.method != 'POST':
        return redirect('landing_page')

    name = request.POST.get('name', '').strip()
    email = request.POST.get('email', '').strip()
    phone = request.POST.get('phone', '').strip()
    message_text = request.POST.get('message', '').strip()

    if not name or not email or not message_text:
        messages.error(request, 'Please complete the contact form.')
        return redirect('landing_page')

    try:
        from .models import ContactMessage
        ContactMessage.objects.create(
            name=name,
            email=email,
            phone=phone,
            message=message_text,
            browser_used=request.META.get('HTTP_USER_AGENT', '')[:500],
            ip_address=request.META.get('REMOTE_ADDR', '')[:45],
        )

        support = getattr(settings, 'SUPPORT_EMAIL', None) or getattr(settings, 'DEFAULT_FROM_EMAIL', None)
        subject = f"Website contact from {name}"
        body = f"Name: {name}\nEmail: {email}\nPhone: {phone}\n\nMessage:\n{message_text}"

        if support:
            sent = send_email(to_email=support, subject=subject, message=body, html=False)
            if not sent:
                messages.warning(request, f'Thank you — your message was saved, but we could not deliver notification to support. Contact: {support}')
            else:
                messages.success(request, 'Thank you — your message was sent. We will get back to you shortly.')
        else:
            messages.success(request, 'Thank you — your message was saved. We will get back to you shortly.')
    except Exception as e:
        messages.error(request, f'Could not send message: {str(e)}')

    return redirect('landing_page')


@login_required
@superuser_required
def superuser_contact_reply(request, message_id):
    message = get_object_or_404(ContactMessage, id=message_id)

    if request.method == 'POST':
        reply_text = (request.POST.get('reply') or '').strip()

        if not reply_text:
            messages.error(request, 'Reply text cannot be empty.')
        else:
            message.reply = reply_text
            message.handled = True
            message.replied_at = timezone.now()
            message.save(update_fields=['reply', 'handled', 'replied_at'])
            messages.success(request, 'Support reply saved.')

    return redirect('superuser_dashboard')


@login_required
def request_voucher(request):

    school = request.user.school

    # compute current term and student count for display
    today = timezone.now().date()
    current_term = Term.objects.filter(
        school=school,
        start_date__lte=today,
        end_date__gte=today
    ).first()

    if not current_term:
        current_term = Term.objects.filter(school=school).order_by("-start_date").first()

    term_name = current_term.name if current_term else ""
    student_count_display = Student.objects.filter(school=school).count()

    if request.method == "POST":

        # Determine current term for the school (prefer active term)
        today = timezone.now().date()
        current_term = Term.objects.filter(
            school=school,
            start_date__lte=today,
            end_date__gte=today
        ).first()

        if not current_term:
            current_term = Term.objects.filter(school=school).order_by("-start_date").first()

        term_name = current_term.name if current_term else ""

        # Count students automatically
        student_count = Student.objects.filter(school=school).count()

        VoucherRequest.objects.create(
            school=school,
            term=term_name,
            student_count=student_count,
            status="pending"
        )

        messages.success(
            request,
            "Voucher request sent successfully"
        )

        return redirect("dos_dashboard")

    # GET REQUEST
    terms = Term.objects.filter(
        school=school
    )

    context = {
        "terms": terms,
        "detected_term": term_name,
        "detected_student_count": student_count_display,
    }

    return render(
        request,
        "dashboards/request_voucher.html",
        context
    )


@login_required
def edit_notice(request, notice_id):
    notice = get_object_or_404(SchoolNotice, id=notice_id, school=request.user.school)

    # Only sender, principal or superuser can edit
    allowed = (request.user == notice.sender) or getattr(request.user, 'role', '') == 'principal' or request.user.is_superuser
    if not allowed:
        messages.error(request, "Permission denied.")
        return redirect('principal_dashboard')

    if request.method == 'POST':
        notice.title = request.POST.get('title', notice.title)
        notice.message = request.POST.get('message', notice.message)
        notice.recipient_type = request.POST.get('recipient_type', notice.recipient_type)
        notice.is_urgent = bool(request.POST.get('is_urgent'))
        notice.save()
        messages.success(request, "Announcement updated.")
        return redirect('principal_dashboard')

    return render(request, 'schools/edit_notice.html', {'notice': notice})


@login_required
def followup_notice(request, notice_id):
    notice = get_object_or_404(SchoolNotice, id=notice_id, school=request.user.school)

    # Allow principals, dos, or sender to follow up
    allowed = getattr(request.user, 'role', '') in ['principal', 'dos'] or request.user == notice.sender or request.user.is_superuser
    if not allowed:
        messages.error(request, "Permission denied.")
        return redirect('principal_dashboard')

    if request.method == 'POST':
        follow_text = request.POST.get('follow_up', '').strip()
        if follow_text:
            notice.follow_up = follow_text
            notice.followed_by = request.user
            notice.followed_at = timezone.now()
            notice.save()
            messages.success(request, "Follow-up saved.")
        else:
            messages.error(request, "Follow-up text cannot be empty.")

    return redirect('principal_dashboard')


@login_required
def delete_notice(request, notice_id):
    notice = get_object_or_404(SchoolNotice, id=notice_id, school=request.user.school)
    allowed = request.user == notice.sender or getattr(request.user, 'role', '') == 'principal' or request.user.is_superuser
    if request.method == 'POST' and allowed:
        notice.delete()
        messages.success(request, "Announcement deleted.")
    else:
        messages.error(request, "Permission denied or invalid request.")
    return redirect('principal_dashboard')

@login_required
def approve_voucher(request, id):

    voucher = get_object_or_404(
        VoucherRequest,
        id=id
    )

    voucher.status = "approved"
    voucher.save()

    messages.success(
        request,
        "Voucher approved successfully"
    )

    return redirect("superuser_dashboard")

@login_required
def superuser_dashboard(request):

    # ----------------------------
    # SCHOOLS
    # ----------------------------
    schools = School.objects.all().order_by("-id")

    active_schools = schools.filter(
        is_active=True
    ).count()

    # ----------------------------
    # APPROVED VOUCHERS
    # ----------------------------
    approved_vouchers = VoucherRequest.objects.filter(
        status="approved"
    ).order_by("-id")

    # ----------------------------
    # PENDING VOUCHERS
    # ----------------------------
    vouchers = (
        VoucherRequest.objects
        .select_related("school")
        .filter(status="pending")
        .order_by("-id")
    )

    pending_vouchers = vouchers.count()

    # ----------------------------
    # DOS MESSAGES
    # ----------------------------
    queries = (
        DOSMessage.objects
        .select_related(
            "school",
            "sender",
            "receiver"
        )
        .filter(receiver=request.user)
        .order_by("-created_at")
    )

    unread_queries = queries.filter(
        status="pending"
    ).count()

    # ----------------------------
    # ADMIN NOTIFICATIONS
    # ----------------------------
    notifications = Notification.objects.filter(recipient=request.user).order_by("-created_at")
    unread_notifications = notifications.filter(is_read=False).count()

    error_reports = ErrorReport.objects.order_by("-created_at")
    unread_error_reports = error_reports.filter(is_read=False).count()
    security_logs = SecurityLog.objects.select_related("user").order_by("-created_at")[:6]
    for log in security_logs:
        log.browser = log.browser or "Unknown"
        log.location = log.location or ("Local environment" if log.ip_address in {"127.0.0.1", "::1", "localhost"} else "Unknown")
    contact_messages = ContactMessage.objects.order_by("-created_at")[:10]

    # ----------------------------
    # CONTEXT
    # ----------------------------
    # include pending license renewals so superusers can see reactivation requests
    from .models import LicenseRenewal
    pending_renewals = LicenseRenewal.objects.filter(status="pending").select_related("school", "requested_by").order_by("-requested_at")
    principals = Principal.objects.select_related("school").all()
    doss = DirectorOfStudies.objects.select_related("school").all()

    context = {

        "schools": schools,

        "active_schools": active_schools,

        "approved_vouchers": approved_vouchers,

        "pending_vouchers": pending_vouchers,

        "vouchers": vouchers,

        "queries": queries,

        "unread_queries": unread_queries,
        "notifications": notifications,
        "unread_notifications": unread_notifications,
        "error_reports": error_reports,
        "unread_error_reports": unread_error_reports,
        "pending_renewals": pending_renewals,
        "pending_renewals_count": pending_renewals.count(),
        "principals": principals,
        "doss": doss,
        "security_logs": security_logs,
        "contact_messages": contact_messages,
    }

    return render(
        request,
        "dashboards/superuser.html",
        context
    )


@login_required
@superuser_required
def security_logs(request):
    logs = SecurityLog.objects.select_related("user").order_by("-created_at")
    for log in logs:
        log.browser = log.browser or "Unknown"
        log.location = log.location or ("Local environment" if log.ip_address in {"127.0.0.1", "::1", "localhost"} else "Unknown")
    return render(request, "dashboards/security_logs.html", {
        "security_logs": logs,
    })


@login_required
@superuser_required
def error_reports(request):
    reports = ErrorReport.objects.order_by("-created_at")
    unread_error_reports = reports.filter(is_read=False).count()
    return render(request, "dashboards/error_reports.html", {
        "reports": reports,
        "unread_error_reports": unread_error_reports,
    })


@login_required
@superuser_required
def mark_error_report_read(request, report_id):
    report = get_object_or_404(ErrorReport, id=report_id)
    if not report.is_read:
        report.is_read = True
        report.save(update_fields=["is_read"])
    referer = request.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    return redirect('error_reports')


def bad_request(request, exception=None):
    return render(request, "exams/errors/400.html", {"message": str(exception)}, status=400)


def not_found(request, exception=None):
    return render(request, "exams/errors/404.html", {"message": str(exception)}, status=404)


def server_error(request):
    return render(request, "exams/errors/500.html", status=500)


def _superuser_management_context():
    schools = School.objects.all().order_by("-id")
    principals = Principal.objects.select_related("school", "user").all().order_by("-id")
    doss = DirectorOfStudies.objects.select_related("school", "user").all().order_by("-id")

    return {
        "schools": schools,
        "principals": principals,
        "doss": doss,
    }


def _process_school_submission(request):
    action = request.POST.get("action", "create")
    school_id = request.POST.get("school_id")
    name = (request.POST.get("name") or "").strip()
    address = (request.POST.get("address") or "").strip()
    phone = (request.POST.get("phone") or "").strip()
    email = (request.POST.get("email") or "").strip()
    subscription_balance = request.POST.get("subscription_balance") or 0
    motto = (request.POST.get("motto") or "").strip()
    logo = request.FILES.get("logo")

    if not all([name, address, phone, email]):
        messages.error(request, "Please fill in all required school fields.")
        return False

    if action == "update":
        school = get_object_or_404(School, id=school_id)

        if School.objects.exclude(id=school.id).filter(name=name).exists():
            messages.warning(request, "Another school already uses that name.")
            return False

        if School.objects.exclude(id=school.id).filter(email=email).exists():
            messages.warning(request, "Another school already uses that email.")
            return False

        if School.objects.exclude(id=school.id).filter(phone=phone).exists():
            messages.warning(request, "Another school already uses that phone number.")
            return False

        school.name = name
        school.address = address
        school.phone = phone
        school.email = email
        school.motto = motto
        school.subscription_balance = subscription_balance

        if logo:
            school.logo = logo

        school.save()
        messages.success(request, f"{school.name} updated successfully.")
        return True

    if School.objects.filter(name=name).exists():
        messages.warning(request, "A school with that name already exists.")
        return False

    if School.objects.filter(email=email).exists():
        messages.warning(request, "A school with that email already exists.")
        return False

    if School.objects.filter(phone=phone).exists():
        messages.warning(request, "A school with that phone number already exists.")
        return False

    School.objects.create(
        name=name,
        address=address,
        phone=phone,
        email=email,
        motto=motto,
        subscription_balance=subscription_balance,
        logo=logo,
        is_active=False,
    )

    messages.success(request, "School added successfully.")
    messages.warning(
        request,
        "New school accounts expire within 48 hours if not activated. Contact admin to activate."
    )
    return True


def _process_principal_submission(request):
    action = request.POST.get("action", "create")
    principal_id = request.POST.get("principal_id")
    name = (request.POST.get("name") or "").strip()
    email = (request.POST.get("email") or "").strip()
    phone = (request.POST.get("phone") or "").strip()
    password = request.POST.get("password") or ""
    school_id = request.POST.get("school")

    if not all([name, email, phone, school_id]):
        messages.error(request, "Please fill all required principal fields.")
        return False

    school = get_object_or_404(School, id=school_id)

    if action == "update":
        principal = get_object_or_404(Principal, id=principal_id)
        user = principal.user

        if user and CustomUser.objects.exclude(id=user.id).filter(email=email).exists():
            messages.warning(request, "Another user already uses that email.")
            return False

        if Principal.objects.exclude(id=principal.id).filter(phone=phone).exists():
            messages.warning(request, "Another principal already uses that phone number.")
            return False

        if Principal.objects.exclude(id=principal.id).filter(school=school).exists():
            messages.warning(request, "This school already has a Principal account.")
            return False

        principal.name = name
        principal.email = email
        principal.phone = phone
        principal.school = school
        principal.save()

        if user:
            user.email = email
            user.school = school
            user.role = "principal"
            user.save(update_fields=["email", "school", "role"])

            if password:
                user.set_password(password)
                user.save(update_fields=["password"])

        messages.success(request, f"{name} updated successfully.")
        return True

    if not password:
        messages.error(request, "Password is required when creating a principal account.")
        return False

    if CustomUser.objects.filter(email=email).exists():
        messages.warning(request, "Email already taken.")
        return False

    if Principal.objects.filter(phone=phone).exists():
        messages.warning(request, "Phone number already used.")
        return False

    if Principal.objects.filter(school=school).exists():
        messages.warning(request, "This school already has a Principal account.")
        return False

    user = CustomUser.objects.create_user(
        email=email,
        password=password,
        role="principal",
        school=school,
        email_verified=False,
    )

    Principal.objects.create(
        user=user,
        school=school,
        name=name,
        email=email,
        phone=phone,
    )

    send_user_verification_email(user, request=request, role_name='Principal')
    messages.success(request, f"{name} registered successfully as Principal. A verification email has been sent.")
    return True


def _process_dos_submission(request):
    action = request.POST.get("action", "create")
    dos_id = request.POST.get("dos_id")
    name = (request.POST.get("name") or "").strip()
    email = (request.POST.get("email") or "").strip()
    phone = (request.POST.get("phone") or "").strip()
    password = request.POST.get("password") or ""
    school_id = request.POST.get("school")

    if not all([name, email, phone, school_id]):
        messages.error(request, "Please fill all required DOS fields.")
        return False

    school = get_object_or_404(School, id=school_id)

    if action == "update":
        dos = get_object_or_404(DirectorOfStudies, id=dos_id)
        user = dos.user

        if user and CustomUser.objects.exclude(id=user.id).filter(email=email).exists():
            messages.warning(request, "Another user already uses that email.")
            return False

        if DirectorOfStudies.objects.exclude(id=dos.id).filter(phone=phone).exists():
            messages.warning(request, "Another DOS already uses that phone number.")
            return False

        if DirectorOfStudies.objects.exclude(id=dos.id).filter(school=school).exists():
            messages.warning(request, "This school already has a DOS account.")
            return False

        dos.name = name
        dos.email = email
        dos.phone = phone
        dos.school = school
        dos.save()

        if user:
            user.email = email
            user.school = school
            user.role = "dos"
            user.save(update_fields=["email", "school", "role"])

            if password:
                user.set_password(password)
                user.save(update_fields=["password"])

        messages.success(request, f"{name} updated successfully.")
        return True

    if not password:
        messages.error(request, "Password is required when creating a DOS account.")
        return False

    if CustomUser.objects.filter(email=email).exists():
        messages.warning(request, "Email already taken.")
        return False

    if DirectorOfStudies.objects.filter(phone=phone).exists():
        messages.warning(request, "Phone number already used.")
        return False

    if DirectorOfStudies.objects.filter(school=school).exists():
        messages.warning(request, "This school already has a DOS account.")
        return False

    user = CustomUser.objects.create_user(
        email=email,
        password=password,
        role="dos",
        school=school,
        email_verified=False,
    )

    DirectorOfStudies.objects.create(
        user=user,
        school=school,
        name=name,
        email=email,
        phone=phone,
    )

    send_user_verification_email(user, request=request, role_name='Director of Studies')
    messages.success(request, f"{name} registered successfully. A verification email has been sent.")
    return True


@superuser_required
def manage_schools(request):
    if request.method == "POST":
        _process_school_submission(request)
        return redirect("manage_schools")

    context = _superuser_management_context()
    return render(request, "dashboards/manage_schools.html", context)


@superuser_required
def manage_principals(request):
    if request.method == "POST":
        _process_principal_submission(request)
        return redirect("manage_principals")

    context = _superuser_management_context()
    return render(request, "dashboards/manage_principals.html", context)


@superuser_required
def edit_principal_by_superuser(request, principal_id):
    principal = get_object_or_404(Principal, id=principal_id)

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        email = (request.POST.get("email") or "").strip()
        phone = (request.POST.get("phone") or "").strip()
        password = request.POST.get("password") or ""
        school_id = request.POST.get("school")

        if not all([name, email, phone, school_id]):
            messages.error(request, "Please fill all required principal fields.")
            return redirect("edit_principal_by_superuser", principal_id=principal.id)

        school = get_object_or_404(School, id=school_id)

        if CustomUser.objects.exclude(id=principal.user_id).filter(email=email).exists():
            messages.warning(request, "Another user already uses that email.")
            return redirect("edit_principal_by_superuser", principal_id=principal.id)

        if Principal.objects.exclude(id=principal.id).filter(phone=phone).exists():
            messages.warning(request, "Another principal already uses that phone number.")
            return redirect("edit_principal_by_superuser", principal_id=principal.id)

        if Principal.objects.exclude(id=principal.id).filter(school=school).exists():
            messages.warning(request, "This school already has a Principal account.")
            return redirect("edit_principal_by_superuser", principal_id=principal.id)

        principal.name = name
        principal.email = email
        principal.phone = phone
        principal.school = school
        principal.save()

        if principal.user:
            principal.user.email = email
            principal.user.school = school
            principal.user.role = "principal"
            principal.user.save(update_fields=["email", "school", "role"])

            if password:
                principal.user.set_password(password)
                principal.user.save(update_fields=["password"])

        messages.success(request, f"{name} updated successfully.")
        return redirect("manage_principals")

    context = _superuser_management_context()
    context.update({"principal": principal})
    return render(request, "dashboards/edit_principal.html", context)


@superuser_required
def delete_principal_by_superuser(request, principal_id):
    principal = get_object_or_404(Principal, id=principal_id)

    if request.method == "POST":
        principal_name = principal.name
        user = principal.user
        principal.delete()
        if user:
            user.delete()
        messages.success(request, f"{principal_name} deleted successfully.")
        return redirect("manage_principals")

    return redirect("manage_principals")


@superuser_required
def manage_dos(request):
    if request.method == "POST":
        _process_dos_submission(request)
        return redirect("manage_dos")

    context = _superuser_management_context()
    return render(request, "dashboards/manage_dos.html", context)


@superuser_required
def edit_dos_by_superuser(request, dos_id):
    dos = get_object_or_404(DirectorOfStudies, id=dos_id)

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        email = (request.POST.get("email") or "").strip()
        phone = (request.POST.get("phone") or "").strip()
        password = request.POST.get("password") or ""
        school_id = request.POST.get("school")

        if not all([name, email, phone, school_id]):
            messages.error(request, "Please fill all required DOS fields.")
            return redirect("edit_dos_by_superuser", dos_id=dos.id)

        school = get_object_or_404(School, id=school_id)

        if CustomUser.objects.exclude(id=dos.user_id).filter(email=email).exists():
            messages.warning(request, "Another user already uses that email.")
            return redirect("edit_dos_by_superuser", dos_id=dos.id)

        if DirectorOfStudies.objects.exclude(id=dos.id).filter(phone=phone).exists():
            messages.warning(request, "Another DOS already uses that phone number.")
            return redirect("edit_dos_by_superuser", dos_id=dos.id)

        if DirectorOfStudies.objects.exclude(id=dos.id).filter(school=school).exists():
            messages.warning(request, "This school already has a DOS account.")
            return redirect("edit_dos_by_superuser", dos_id=dos.id)

        dos.name = name
        dos.email = email
        dos.phone = phone
        dos.school = school
        dos.save()

        if dos.user:
            dos.user.email = email
            dos.user.school = school
            dos.user.role = "dos"
            dos.user.save(update_fields=["email", "school", "role"])

            if password:
                dos.user.set_password(password)
                dos.user.save(update_fields=["password"])

        messages.success(request, f"{name} updated successfully.")
        return redirect("manage_dos")

    context = _superuser_management_context()
    context.update({"dos": dos})
    return render(request, "dashboards/edit_dos.html", context)


@superuser_required
def delete_dos_by_superuser(request, dos_id):
    dos = get_object_or_404(DirectorOfStudies, id=dos_id)

    if request.method == "POST":
        dos_name = dos.name
        user = dos.user
        dos.delete()
        if user:
            user.delete()
        messages.success(request, f"{dos_name} deleted successfully.")
        return redirect("manage_dos")

    return redirect("manage_dos")
from django.contrib import messages
from .models import DOSQuery
from django.shortcuts import redirect
from django.contrib import messages
from .models import DOSMessage
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.shortcuts import redirect

User = get_user_model()


def send_query(request):

    if request.method == "POST":

        subject = request.POST.get("subject")
        message_text = request.POST.get("message")

        superuser = User.objects.filter(
            is_superuser=True
        ).first()

        if not superuser:

            messages.error(
                request,
                "No superuser account found"
            )

            return redirect("dos_dashboard")

        DOSMessage.objects.create(
            school=request.user.school,
            sender=request.user,
            receiver=superuser,
            subject=subject,
            message=message_text,
            parent=None,
            status="pending"
        )

        try:
            if settings.EMAIL_HOST_USER and superuser.email:
                send_email(
                    to_email=[superuser.email],
                    subject=f"DOS query: {subject}",
                    message=f"{request.user.get_full_name() or request.user.email} sent a DOS query:\n\n{message_text}",
                    recipient_name=superuser.get_full_name() or superuser.email,
                    html=False,
                )
        except Exception:
            pass

        messages.success(
            request,
            "Message sent successfully"
        )

    return redirect("dos_dashboard")

@login_required
def reply_query(request, query_id):

    query = DOSMessage.objects.get(id=query_id)

    if request.method == "POST":

        reply_message = request.POST.get("reply")

        # =========================
        # CREATE REPLY
        # =========================
        DOSMessage.objects.create(

            school=query.school,

            sender=request.user,

            receiver=query.sender,

            subject=f"RE: {query.subject}",

            message=reply_message,

            parent=query,

            status="pending"
        )

        # =========================
        # UPDATE ORIGINAL MESSAGE
        # =========================
        query.status = "replied"
        query.save()

        messages.success(
            request,
            "Reply sent successfully"
        )

        return redirect("superuser_dashboard")

def send_dos_message(request):
    if request.method == "POST":
        DOSMessage.objects.create(
            school=request.user.school,
            sender=request.user,
            receiver_id=request.POST.get("receiver"),
            subject=request.POST.get("subject"),
            message=request.POST.get("message"),
        )
        return redirect("dos_dashboard")
    
from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import DOSMessage

@login_required
def reply_message(request, id):
    parent = get_object_or_404(DOSMessage, id=id)

    if request.method == "POST":
        reply_text = request.POST.get("message", "").strip()
        if not reply_text:
            messages.error(request, "Reply cannot be empty.")
            return redirect("dos_dashboard")

        # create reply
        reply = DOSMessage.objects.create(
            school=parent.school,
            sender=request.user,
            receiver=parent.sender,
            subject="RE: " + (parent.subject or "Message"),
            message=reply_text,
            parent=parent,
            status="pending"  # reply itself starts as pending for the receiver
        )

        # mark parent as replied and clear pending/unread flags for the DOS
        parent.status = "replied"
        parent.save(update_fields=["status"])

        # Optionally mark any pending replies to this parent as not unread
        # (depends on how you track unread; adjust to your model)
        DOSMessage.objects.filter(parent=parent, receiver=request.user, status="pending").update(status="read")

        messages.success(request, "Reply sent.")
        return redirect("dos_dashboard")

    # If GET, redirect back (or render a reply form if you prefer)
    return redirect("dos_dashboard")
@login_required
def clear_replied_count(request):
    DOSMessage.objects.filter(receiver=request.user, status="replied").update(status="cleared")
    messages.success(request, "Replied message count cleared.")
    return redirect("dos_dashboard")

@login_required
def clear_all_messages(request):
    """Clear all pending messages/queries for superuser and DOS"""
    if request.method == "POST":
        user = request.user
        # Mark all DOS messages as cleared
        DOSMessage.objects.filter(receiver=user).update(status="cleared")
        messages.success(request, "✓ All messages cleared successfully!")
        
        # Redirect to appropriate dashboard
        if user.is_superuser:
            return redirect("superuser_dashboard")
        else:
            return redirect("dos_dashboard")
    
    # GET request - show confirmation
    return redirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def delete_single_message(request, message_id):
    """Delete a single message"""
    if request.method == "POST":
        try:
            message = DOSMessage.objects.get(id=message_id)
            if message.receiver == request.user or request.user.is_superuser:
                message.delete()
                messages.success(request, "Message deleted successfully!")
        except DOSMessage.DoesNotExist:
            messages.error(request, "Message not found.")
    
    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
def delete_message_completely(request, message_id):
    """Permanently delete a message and its replies."""
    if request.method == "POST":
        try:
            msg = DOSMessage.objects.get(id=message_id)
            # Only allow receiver or superuser to purge
            if msg.receiver == request.user or request.user.is_superuser:
                # delete children explicitly (FK CASCADE should handle it)
                DOSMessage.objects.filter(parent=msg).delete()
                msg.delete()
                messages.success(request, "Message permanently deleted.")
            else:
                messages.error(request, "Permission denied.")
        except DOSMessage.DoesNotExist:
            messages.error(request, "Message not found.")

    return redirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def delete_message(request, message_id):
    """Delete a single message."""
    if request.method == 'POST':
        try:
            message = DOSMessage.objects.get(id=message_id)
            if message.receiver == request.user or request.user.is_superuser:
                message.delete()
                messages.success(request, "Message deleted successfully.")
            else:
                messages.error(request, "Permission denied.")
        except DOSMessage.DoesNotExist:
            messages.error(request, "Message not found.")

    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
def mark_notification_read(request, notification_id):
    """Mark a single notification as read."""
    notification = get_object_or_404(Notification, id=notification_id, recipient=request.user)
    notification.is_read = True
    notification.save()
    return redirect(request.META.get('HTTP_REFERER', 'dos_dashboard'))


@login_required
def mark_all_notifications_read(request):
    """Mark all notifications as read."""
    Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    messages.success(request, "All notifications have been marked as read.")
    return redirect(request.META.get('HTTP_REFERER', 'dos_dashboard'))


@login_required
def send_school_notice(request):
    """Send a notice to teachers, students, or both."""
    school = getattr(request.user, 'school', None)
    if school is None:
        messages.error(request, "You are not assigned to a school.")
        return redirect('landing_page')

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        message_text = request.POST.get('message', '').strip()
        recipient_type = request.POST.get('recipient_type', 'all')
        is_urgent = bool(request.POST.get('is_urgent'))

        if not title or not message_text:
            messages.error(request, "Title and message are required.")
            return render(request, 'schools/send_notice.html', {'school': school})

        SchoolNotice.objects.create(
            school=school,
            sender=request.user,
            title=title,
            message=message_text,
            recipient_type=recipient_type,
            is_urgent=is_urgent,
        )

        messages.success(request, "School notice sent successfully.")
        return redirect('principal_dashboard')

    return render(request, 'schools/send_notice.html', {'school': school})


@login_required
def edit_school_info(request):
    print(">>> edit_school_info() CALLED <<<")
    school = getattr(request.user, 'school', None)
    if school is None:
        messages.error(request, "You are not assigned to a school.")
        return redirect('landing_page')

    #normalize_school_logo_field(school)# type: ignore

    if request.method == 'POST':
        school.name = request.POST.get('name', school.name)
        school.motto = request.POST.get('motto', school.motto)
        school.address = request.POST.get('address', school.address)
        school.email = request.POST.get('email', school.email)
        school.phone = request.POST.get('phone', school.phone)
        school.bank_name = request.POST.get('bank_name', school.bank_name)
        school.account_number = request.POST.get('account_number', school.account_number)

        logo = request.FILES.get("logo")
        if logo:
            print("=" * 60)
            print("UPLOADED FILE")
            print("Uploaded filename :", logo.name)
            print("Uploaded file type:", type(logo))

            school.logo = logo

            print("\nAFTER ASSIGNMENT")
            print("school.logo      :", school.logo)
            print("school.logo.name :", getattr(school.logo, 'name', None))
            print("school.logo.public_id :", getattr(school.logo, 'public_id', None))
            try:
                print("school.logo.url  :", school.logo.url)
            except Exception as e:
                print("URL not available:", e)

        print("\nCALLING school.save()...")
        school.save()

        print("\nAFTER school.save()")
        print("school.logo      :", school.logo)
        print("school.logo.name :", getattr(school.logo, 'name', None))
        print("school.logo.public_id :", getattr(school.logo, 'public_id', None))
        try:
            print("school.logo.url  :", school.logo.url)
        except Exception as e:
            print("URL error:", e)

        print("=" * 60)

        try:
            print("school.logo.url  :", school.logo.url)
        except Exception as e:
            print("URL error:", e)

        print("=" * 60)

        school.save()
        messages.success(request, "School information updated successfully.")
        return redirect('principal_dashboard')

    return render(request, 'schools/edit_school_info.html', {'school': school})


@login_required
def principal_school_manager(request):
    """Principal page for editing school info and managing the school DOS account."""
    school = getattr(request.user, 'school', None)
    if school is None:
        messages.error(request, "You are not assigned to a school.")
        return redirect('landing_page')

    dos = getattr(school, 'dos', None)
    #normalize_school_logo_field(school) # type: ignore

    if request.method == 'POST':
        section = request.POST.get('section')

        if section == 'school':
            school.name = request.POST.get('name', school.name)
            school.motto = request.POST.get('motto', school.motto)
            school.address = request.POST.get('address', school.address)
            school.email = request.POST.get('email', school.email)
            school.phone = request.POST.get('phone', school.phone)

            logo = request.FILES.get('logo')

            if logo:
                print("UPLOADED FILE")
                print("Uploaded filename :", logo.name)
                print("Uploaded file type:", type(logo))

                school.logo = logo

                print("\nAFTER ASSIGNMENT")
                print("school.logo      :", school.logo)
                print("school.logo.name :", getattr(school.logo, 'name', None))
                print("school.logo.public_id :", getattr(school.logo, 'public_id', None))

                try:
                    print("school.logo.url  :", school.logo.url)
                except Exception as e:
                    print("URL not available:", e)

            print("\nCALLING school.save()...")
            school.save()

            print("\nAFTER school.save()")
            print("school.logo      :", school.logo)
            print("school.logo.name :", getattr(school.logo, 'name', None))
            print("school.logo.public_id :", getattr(school.logo, 'public_id', None))

            try:
                print("school.logo.url  :", school.logo.url)
            except Exception as e:
                print("URL error:", e)

            print("=" * 60)

            messages.success(request, "School information updated successfully.")
            return redirect('principal_school_manager')

        if section == 'dos':
            name = (request.POST.get('name') or '').strip()
            email = (request.POST.get('email') or '').strip()
            phone = (request.POST.get('phone') or '').strip()
            password = request.POST.get('password') or ''

            if not all([name, email, phone]):
                messages.error(request, "Please fill all DOS fields.")
                return redirect('principal_school_manager')

            if dos:
                user = dos.user
                if user and CustomUser.objects.exclude(id=user.id).filter(email=email).exists():
                    messages.warning(request, "Another user already uses that email.")
                    return redirect('principal_school_manager')
                if DirectorOfStudies.objects.exclude(id=dos.id).filter(phone=phone).exists():
                    messages.warning(request, "Another DOS already uses that phone number.")
                    return redirect('principal_school_manager')

                dos.name = name
                dos.email = email
                dos.phone = phone
                dos.save()

                if user:
                    user.email = email
                    user.school = school
                    user.role = 'dos'
                    user.save(update_fields=['email', 'school', 'role'])
                    if password:
                        user.set_password(password)
                        user.save(update_fields=['password'])

                messages.success(request, "DOS information updated successfully.")
                return redirect('principal_school_manager')

            if not password:
                messages.error(request, "Password is required when creating a DOS account.")
                return redirect('principal_school_manager')

            if CustomUser.objects.filter(email=email).exists():
                messages.warning(request, "Email already taken.")
                return redirect('principal_school_manager')

            if DirectorOfStudies.objects.filter(phone=phone).exists():
                messages.warning(request, "Phone number already used.")
                return redirect('principal_school_manager')

            user = CustomUser.objects.create_user(
                email=email,
                password=password,
                role='dos',
                school=school,
                email_verified=False,
            )

            DirectorOfStudies.objects.create(
                user=user,
                school=school,
                name=name,
                email=email,
                phone=phone,
            )
            send_user_verification_email(user, request=request, role_name='Director of Studies')
            messages.success(request, "DOS account created successfully. A verification email has been sent.")
            return redirect('principal_school_manager')

    return render(request, 'schools/principal_school_manager.html', {
        'school': school,
        'dos': dos,
    })


@login_required
def reset_notification_count(request):
    """Reset notification counts to zero"""
    if request.method == "POST":
        user = request.user
        # Mark all unread/pending messages as read
        DOSMessage.objects.filter(
            receiver=user,
            status__in=['pending', 'new']
        ).update(status="cleared")
        
        messages.success(request, "✓ Notifications reset to zero!")
        
        if user.is_superuser:
            return redirect("superuser_dashboard")
        else:
            return redirect("dos_dashboard")
    
    return redirect(request.META.get('HTTP_REFERER', '/'))

    
def landing_page(request):
    # Show verified schools carousel and testimonials on landing
    schools_list = School.objects.filter(is_verified=True).order_by('-created_at')[:12]
    
    # Simple static testimonials; could be replaced by a model later
    testimonials = [
        {"author": "St. Mary's High", "text": "Brainet transformed our reporting and saved hours each week."},
        {"author": "Green Valley Academy", "text": "Reliable, fast and great support."},
        {"author": "Sunrise School", "text": "Teachers love the online class features and easy grading."},
    ]
    return render(request, "dashboards/landing.html", {"schools_list": schools_list, "testimonials": testimonials})


def mobile_app(request):
    return render(request, "dashboards/mobile_app.html")


@login_required
def features_demo(request):
    """Demo page showcasing new features for customers"""
    return render(request, "schools/features_demo.html")


def about_us(request):
    """About page with developer information"""
    return render(request, "schools/about_us.html")


def contact_admin(request):
    """Simple page instructing users to contact their administrator."""
    return render(request, "schools/contact_admin.html")

def view_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    return render(request, "dos/view_students.html", {"student": student})


@login_required
def dos_dashboard(request):
    school = request.user.school

    # =========================
    # BASIC DATA
    # =========================
    classes = Class.objects.filter(school=school)
    students = Student.objects.filter(school=school)
    subjects = Subject.objects.filter(school=school)
    exams = Exam.objects.filter(school=school).order_by("-created_at")

    exam_window_open, active_exam_count = resolve_exam_window_state(request, school)

    # =========================
    # DOS MESSAGES (both sent and received)
    # =========================
    dos_messages = DOSMessage.objects.filter(
        school=school,
        parent__isnull=True
    ).order_by("-created_at")

    # only unread replies to DOS
    unread_messages = DOSMessage.objects.filter(
        school=school,
        receiver=request.user,
        status="pending"
    ).count()

    # =========================
    # VOUCHER REQUESTS
    # =========================
    vouchers = VoucherRequest.objects.filter(school=school).order_by("-id")
    pending_vouchers = vouchers.filter(status="pending").count()
    approved_vouchers = vouchers.filter(status="approved").count()
    rejected_vouchers = vouchers.filter(status="rejected").count()

    # =========================
    # CONTEXT
    # =========================
    context = {
        "school": school,
        "classes": classes,
        "students": students,
        "subjects": subjects,
        "exams": exams,
        "dos_messages": dos_messages,
        "unread_messages": unread_messages,
        "vouchers": vouchers,
        "pending_vouchers": pending_vouchers,
        "approved_vouchers": approved_vouchers,
        "rejected_vouchers": rejected_vouchers,
        "exam_window_open": exam_window_open,
        "active_exam_count": active_exam_count,
        # Notices for staff
        "notices": SchoolNotice.objects.filter(school=school).filter(recipient_type__in=['teachers','all']).order_by('-created_at'),
    }

    return render(request, "dashboards/dos.html", context)
from django.db.models import Avg, Count
@login_required
def principal_dashboard(request):
    school = request.user.school

    # Counts
    student_count = Student.objects.filter(school=school).count()
    teacher_count = Teacher.objects.filter(school=school).count()
    dorm_count = Dormitory.objects.filter(school=school).count()

    # Performance: average marks per subject
    subject_performance = (
        Mark.objects.filter(student__school=school)
        .values("subject__name")
        .annotate(avg_score=Avg("score"), student_count=Count("student"))
        .order_by("subject__name")
    )

    exam_window_open, active_exam_count = resolve_exam_window_state(request, school)

    return render(request, "dashboards/principal.html", {
        "school": school,
        "student_count": student_count,
        "teacher_count": teacher_count,
        "dorm_count": dorm_count,
        "subject_performance": subject_performance,
        "notices_sent": SchoolNotice.objects.filter(school=school).order_by('-created_at'),
        "notices": SchoolNotice.objects.filter(school=school).order_by('-created_at'),
        "exam_window_open": exam_window_open,
        "active_exam_count": active_exam_count,
    })

@login_required
def class_list_selector(request):
    """Select class and stream to export/preview class list"""
    school = request.user.school
    
    classes = Class.objects.filter(school=school).order_by("level", "name")
    teacher_restricted = False
    teacher_assignments = None
    assigned_class_ids = []
    
    teacher = Teacher.objects.filter(user=request.user, school=school).first()
    if teacher:
        teacher_assignments = ClassTeacherAssignment.objects.filter(
            teacher=teacher
        ).select_related("class_obj", "stream")
        if teacher_assignments.exists():
            teacher_restricted = True
            assigned_class_ids = list(teacher_assignments.values_list("class_obj_id", flat=True))
            classes = classes.filter(id__in=assigned_class_ids)
    
    streams = []
    selected_class = None
    
    if request.GET.get("class_id"):
        selected_class_id = request.GET.get("class_id")
        selected_class_qs = Class.objects.filter(id=selected_class_id, school=school)
        if teacher_restricted:
            selected_class_qs = selected_class_qs.filter(id__in=assigned_class_ids)
        selected_class = get_object_or_404(selected_class_qs)
        streams = Stream.objects.filter(class_group=selected_class).order_by("name")
        if teacher_restricted and teacher_assignments is not None:
            valid_stream_ids = list(
                teacher_assignments.filter(class_obj=selected_class).values_list("stream_id", flat=True)
            )
            streams = streams.filter(id__in=valid_stream_ids)
    
    return render(request, "dos/class_list_selector.html", {
        "school": school,
        "classes": classes,
        "streams": streams,
        "selected_class": selected_class,
        "teacher_restricted": teacher_restricted,
    })
    
from django.db.models import Count

@login_required
def manage_classes(request):
    school = request.user.school

    classes = (
        Class.objects
        .filter(school=school)
        .annotate(
            student_count=Count("students"),
            stream_count=Count("streams")
        )
        .prefetch_related("streams")
        .order_by("level")
    )

    teachers = Teacher.objects.filter(school=school)
    subjects = Subject.objects.filter(school=school)

    return render(
        request,
        "dos/classes.html",
        {
            "classes": classes,
            "teachers": teachers,
            "subjects": subjects,
        }
    )


@login_required
def assign_class_master(request):
    """Assign a teacher as class master to a class or stream"""
    if request.method == "POST":
        school = request.user.school
        class_id = request.POST.get("class_id")
        stream_id = request.POST.get("stream_id")
        teacher_id = request.POST.get("class_master")

        class_obj = get_object_or_404(Class, id=class_id, school=school)

        if stream_id:
            # Assign to specific stream
            stream = get_object_or_404(Stream, id=stream_id, class_group=class_obj)
            # Update ClassTeacherAssignment for this stream
            if teacher_id:
                teacher = get_object_or_404(Teacher, id=teacher_id, school=school)
                ClassTeacherAssignment.objects.update_or_create(
                    class_obj=class_obj,
                    stream=stream,
                    defaults={"teacher": teacher}
                )
                messages.success(request, f"Stream master assigned successfully.")
            else:
                ClassTeacherAssignment.objects.filter(
                    class_obj=class_obj,
                    stream=stream
                ).delete()
                messages.success(request, "Stream master removed.")
        else:
            # Assign to entire class
            if teacher_id:
                teacher = get_object_or_404(Teacher, id=teacher_id, school=school)
                class_obj.class_master = teacher
                class_obj.save()
                messages.success(request, f"Class master assigned successfully.")
            else:
                class_obj.class_master = None
                class_obj.save()
                messages.success(request, "Class master removed.")

    return redirect("manage_classes")


@login_required
def add_class_subject(request):
    """Add subjects to a class"""
    if request.method == "POST":
        school = request.user.school
        class_id = request.POST.get("class_id")
        subject_ids = request.POST.getlist("subject_id")

        class_obj = get_object_or_404(Class, id=class_id, school=school)

        # Get all currently assigned subjects
        current_subjects = set(
            ClassSubject.objects.filter(class_name=class_obj).values_list("subject_id", flat=True)
        )

        new_subject_ids = set(int(sid) for sid in subject_ids if sid)

        # Remove subjects not in the new list
        to_remove = current_subjects - new_subject_ids
        if to_remove:
            ClassSubject.objects.filter(class_name=class_obj, subject_id__in=to_remove).delete()

        # Add new subjects
        for subject_id in new_subject_ids - current_subjects:
            subject = get_object_or_404(Subject, id=subject_id, school=school)
            ClassSubject.objects.get_or_create(
                school=school,
                class_name=class_obj,
                subject=subject
            )

        messages.success(request, "Subjects assigned successfully.")

    return redirect("manage_classes")


@login_required
def add_stream(request):
    """Add a stream/division to a class"""
    if request.method == "POST":
        school = request.user.school
        class_id = request.POST.get("class_id")
        stream_name = request.POST.get("stream_name", "").strip()

        class_obj = get_object_or_404(Class, id=class_id, school=school)

        if stream_name:
            # Check if stream already exists
            existing = Stream.objects.filter(
                class_group=class_obj,
                name=stream_name
            ).first()

            if existing:
                messages.warning(request, f"Stream '{stream_name}' already exists for this class.")
            else:
                Stream.objects.create(class_group=class_obj, name=stream_name)
                messages.success(request, f"Stream '{stream_name}' added successfully.")
        else:
            messages.error(request, "Stream name cannot be empty.")

    return redirect("manage_classes")
    
    
@login_required
def add_class(request):
    if request.method == "POST":
        name = request.POST.get("name")
        level = request.POST.get("level")
        stream = request.POST.get("stream")  # <-- capture stream
        # Save class and optional stream
        school = request.user.school

        new_class = Class.objects.create(
            name=name,
            level=level,
            school=school
        )

        # create a Stream if provided
        if stream:
            from classes.models import Stream
            Stream.objects.create(class_group=new_class, name=stream)

        messages.success(request, f"Class {name} created successfully!")
        return redirect("manage_classes")
    return render(request, "dos/add_class.html")

def edit_class(request, class_id):
    c = get_object_or_404(Class, id=class_id)
    teachers = Teacher.objects.all()

    if request.method == "POST":
        c.name = request.POST.get("name")
        c.level = request.POST.get("level")
        c.stream = request.POST.get("stream")
        teacher_id = request.POST.get("teacher")
        c.teacher_id = teacher_id if teacher_id else None
        c.save()
        messages.success(request, "Class updated successfully!")
        return redirect("manage_classes")

    # Pass class object + teachers to template
    return render(request, "dos/edit_class.html", {
        "class_obj": c,
        "teachers": teachers
    })

@login_required
def delete_class(request, class_id):
    school = request.user.school

    school_class = get_object_or_404(
        Class,
        id=class_id,
        school=school
    )

    if request.method == 'POST':
        school_class.delete()
        messages.success(request, "Class deleted successfully.")
        return redirect("manage_classes")

    # Only allow POST deletes
    return redirect("manage_classes")

@login_required
@login_required
def view_class_students(request, class_id):
    """
    View students in a class with optional stream selection.
    Allows DOS/class teacher to select and view students, then export to PDF.
    """
    school = request.user.school
    
    school_class = get_object_or_404(
        Class,
        id=class_id,
        school=school
    )
    
    # Get all streams for this class
    streams = Stream.objects.filter(class_group=school_class).order_by("name")
    
    # Get selected stream if provided
    stream_id = request.GET.get("stream_id")
    selected_stream = None
    
    if stream_id:
        selected_stream = get_object_or_404(Stream, id=stream_id, class_group=school_class)

    # Restrict class teachers to their assigned stream(s) for this class
    teacher_assignments = ClassTeacherAssignment.objects.filter(
        teacher__user=request.user,
        class_obj=school_class
    ).select_related("stream")
    teacher_restricted = teacher_assignments.exists()
    if teacher_restricted:
        valid_stream_ids = list(teacher_assignments.values_list("stream_id", flat=True))
        streams = streams.filter(id__in=valid_stream_ids)
        if not selected_stream or selected_stream.id not in valid_stream_ids:
            selected_stream = Stream.objects.filter(id=valid_stream_ids[0]).first()
    
    # Get students based on selection
    student_query = Student.objects.filter(
        school=school,
        current_class=school_class
    )
    
    # Filter by stream if selected
    if selected_stream:
        students = student_query.filter(stream=selected_stream).order_by("name")
        display_title = f"{school_class.name} - {selected_stream.name}"
    else:
        # If no stream selected, show all unstreamed students
        students = student_query.filter(stream__isnull=True).order_by("name")
        display_title = f"{school_class.name}"
    
    # Count statistics
    total_students = students.count()
    male_count = students.filter(gender__iexact="Male").count()
    female_count = students.filter(gender__iexact="Female").count()
    
    return render(request, "dos/view_class_students.html", {
        "school": school,
        "school_class": school_class,
        "streams": streams,
        "selected_stream": selected_stream,
        "students": students,
        "display_title": display_title,
        "total_students": total_students,
        "male_count": male_count,
        "female_count": female_count,
        "teacher_restricted": teacher_restricted,
    })
    
@login_required 
def manage_students(request):
    school = request.user.school

    students = Student.objects.filter(school=school).order_by("name")

    classes = Class.objects.filter(school=school)
    streams = Stream.objects.filter(class_group__school=school)
    dorms = Dormitory.objects.filter(school=school)
    return render(request, "dos/manage_students.html", {
        "students": students,
        "classes": classes,
        "streams": streams,
        "dorms": dorms,
    })


@login_required
def activate_student(request, student_id):
    student = get_object_or_404(Student, id=student_id, school=request.user.school)
    if request.method == "POST":
        student.status = "active"
        student.current_class = student.current_class or Class.objects.filter(school=request.user.school).order_by("level").first()
        student.save(update_fields=["status", "current_class"])
        messages.success(request, f"{student.name} has been activated.")
    return redirect("manage_students")


@login_required
def deactivate_student(request, student_id):
    student = get_object_or_404(Student, id=student_id, school=request.user.school)
    if request.method == "POST":
        student.status = "inactive"
        student.current_class = None
        student.stream = None
        student.save(update_fields=["status", "current_class", "stream"])
        messages.success(request, f"{student.name} has been deactivated from active school activities.")
    return redirect("manage_students")
from students.models import Student
from django.contrib.auth import get_user_model

User = get_user_model()
@login_required
def get_class_streams(request):
    """API endpoint: Return streams as JSON for a selected class"""
    import json
    from django.http import JsonResponse
    
    class_id = request.GET.get('class_id')
    school = request.user.school
    
    if not class_id:
        return JsonResponse({'streams': []})
    
    try:
        class_obj = Class.objects.get(id=class_id, school=school)
        streams = Stream.objects.filter(class_group=class_obj).values('id', 'name')
        return JsonResponse({
            'streams': list(streams),
            'success': True
        })
    except Class.DoesNotExist:
        return JsonResponse({'streams': [], 'success': False})

@login_required
def add_student(request):

    school = request.user.school

    classes = Class.objects.filter(school=school)
    dormitory = Dormitory.objects.filter(school=school)

    if request.method == "POST":

        name = request.POST.get("name")
        admission_number = request.POST.get("admission_number")
        gender = request.POST.get("gender")
        class_id = request.POST.get("class_id")
        stream_id = request.POST.get("stream_id")
        dorm_id = request.POST.get("dorm_id")

        # =========================
        # CLASS
        # =========================
        school_class = get_object_or_404(Class, id=class_id, school=school)

        # =========================
        # STREAM (OPTIONAL)
        # =========================
        stream = None
        if stream_id:
            stream = get_object_or_404(Stream, id=stream_id, class_group=school_class)

        # =========================
        # DORM (OPTIONAL)
        # =========================
        dorm = None
        if dorm_id:
            dorm = get_object_or_404(Dormitory, id=dorm_id, school=school)

        # =========================
        # EMAIL LOGIN SYSTEM
        # =========================
        email = student_login_email(admission_number, school)

        if User.objects.filter(email=email).exists():
            messages.error(request, "Student already exists")
            return redirect("add_student")

        # =========================
        # CREATE USER
        # =========================
        user = User.objects.create_user(
            email=email,
            password=admission_number,
            school=school,
            role="student"
        )

        # =========================
        # CREATE STUDENT PROFILE
        # =========================
        Student.objects.create(
            user=user,
            school=school,
            name=name,
            admission_number=admission_number,
            gender=gender,
            current_class=school_class,
            stream=stream,
            dormitory=dorm
        )

        messages.success(
            request,
            f"Student created successfully! Login: {email} / {admission_number}"
        )

        return redirect("manage_students")

    return render(request, "dos/add_student.html", {
        "classes": classes,
        "dorms": dormitory,
    })

@login_required
def edit_student(request, student_id):
    school = request.user.school

    student = get_object_or_404(
        Student,
        id=student_id,
        school=school
    )

    if request.method == "POST":
        student.name = request.POST.get("name")
        student.admission_number = request.POST.get("admission_number")
        student.gender = request.POST.get("gender")

        # ======================
        # CLASS
        # ======================
        class_id = request.POST.get("class_id")
        student.current_class = get_object_or_404(
            Class,
            id=class_id,
            school=school
        )

        # ======================
        # STREAM (OPTIONAL)
        # ======================
        stream_id = request.POST.get("stream_id")
        if stream_id:
            student.stream = get_object_or_404(
                Stream,
                id=stream_id,
                class_group=student.current_class
            )
        else:
            student.stream = None

        # ======================
        # DORMITORY (OPTIONAL)
        # ======================
        dorm_id = request.POST.get("dormitory_id")
        if dorm_id:
            student.dormitory = get_object_or_404(
                Dormitory,
                id=dorm_id,
                school=school
            )
        else:
            student.dormitory = None

        student.save()
        messages.success(request, "Student updated successfully.")
        return redirect("manage_students")

    # for form dropdowns
    classes = Class.objects.filter(school=school)
    streams = Stream.objects.filter(class_group__school=school)
    dorms = Dormitory.objects.filter(school=school)

    return render(request, "dos/edit_student.html", {
        "student": student,
        "classes": classes,
        "streams": streams,
        "dorms": dorms,
    })

def delete_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)

    if request.method == "POST":
        student.delete()

    return redirect("manage_students")
    
  
def download_student_list(request):
    school = request.user.school

    students = Student.objects.filter(school=school)

    return HttpResponse("Student list download will be implemented here.")  


@login_required
def download_student_import_template(request):
    school = request.user.school
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Students"
    headers = ["name", "admission_number", "gender", "class_name", "stream_name", "parent_phone"]
    sheet.append(headers)
    sheet.append(["John Doe", "ADM001", "Male", "Grade 1", "A", "0712345678"])
    sheet.append(["Jane Doe", "ADM002", "Female", "Grade 1", "B", "0712345679"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename={school.name.replace(' ', '_')}_student_import_template.xlsx"
    return response


@login_required
def import_students_from_excel(request):
    school = request.user.school

    if request.method != "POST":
        return redirect("manage_students")

    excel_file = request.FILES.get("excel_file")
    class_id = request.POST.get("class_id")
    if not excel_file or not class_id:
        messages.error(request, "Please upload an Excel file and select a class.")
        return redirect("manage_students")

    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception as exc:
        messages.error(request, f"Unable to read Excel file: {exc}")
        return redirect("manage_students")

    if not rows:
        messages.error(request, "The uploaded Excel file is empty.")
        return redirect("manage_students")

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    expected_headers = ["name", "admission_number", "gender", "class_name", "stream_name", "parent_phone"]
    if headers[:len(expected_headers)] != expected_headers:
        messages.error(request, "The Excel template is invalid. Please use the downloaded template.")
        return redirect("manage_students")

    school_class = get_object_or_404(Class, id=class_id, school=school)
    imported = 0
    skipped = 0
    errors = []

    for row_index, row in enumerate(rows[1:], start=2):
        values = [cell if cell is not None else "" for cell in row]
        if not any(str(value).strip() for value in values):
            continue

        name = str(values[0]).strip() if len(values) > 0 else ""
        admission_number = str(values[1]).strip() if len(values) > 1 else ""
        gender = str(values[2]).strip() if len(values) > 2 else ""
        class_name = str(values[3]).strip() if len(values) > 3 else ""
        stream_name = str(values[4]).strip() if len(values) > 4 else ""
        parent_phone = str(values[5]).strip() if len(values) > 5 else ""

        if not name or not admission_number or not gender:
            skipped += 1
            errors.append(f"Row {row_index}: name, admission_number, and gender are required.")
            continue

        if Student.objects.filter(school=school, admission_number=admission_number).exists():
            skipped += 1
            errors.append(f"Row {row_index}: admission number {admission_number} already exists.")
            continue

        if gender not in ["Male", "Female", "male", "female"]:
            skipped += 1
            errors.append(f"Row {row_index}: gender must be Male or Female.")
            continue

        target_class = school_class
        if class_name and class_name != school_class.name:
            target_class = Class.objects.filter(school=school, name__iexact=class_name).first()
            if not target_class:
                skipped += 1
                errors.append(f"Row {row_index}: class {class_name} was not found.")
                continue

        stream = None
        if stream_name:
            stream = Stream.objects.filter(class_group=target_class, name__iexact=stream_name).first()
            if not stream:
                stream = Stream.objects.create(class_group=target_class, name=stream_name)
        elif target_class and target_class.streams.exists():
            default_stream = target_class.streams.first()
            if default_stream:
                stream = default_stream

        email = student_login_email(admission_number, school)
        user = User.objects.create_user(
            email=email,
            password=admission_number,
            school=school,
            role="student"
        )

        Student.objects.create(
            user=user,
            school=school,
            name=name,
            admission_number=admission_number,
            gender=gender.title(),
            current_class=target_class,
            stream=stream,
            parent_phone=parent_phone,
            status="active"
        )
        imported += 1

    if imported:
        messages.success(request, f"Imported {imported} student(s) successfully.")
    if skipped:
        messages.warning(request, f"Skipped {skipped} row(s).")
    if errors:
        messages.info(request, "Import issues: " + " | ".join(errors[:8]))

    return redirect("manage_students")


@login_required
def export_students_to_excel(request):
    school = request.user.school
    students = Student.objects.filter(school=school).order_by("name")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Students"
    headers = ["name", "admission_number", "gender", "class_name", "stream_name", "parent_phone", "status"]
    sheet.append(headers)

    for student in students:
        sheet.append([
            student.name,
            student.admission_number,
            student.gender,
            student.current_class.name if student.current_class else "",
            student.stream.name if student.stream else "",
            student.parent_phone or "",
            student.status,
        ])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename={school.name.replace(' ', '_')}_students_export.xlsx"
    return response


@login_required
def manage_dorms(request):
    school = request.user.school

    return render(request, "dos/dorms.html", {
        "dorms": Dormitory.objects.filter(school=school)
    })
    
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required

from students.models import Student
from .models import StudentMark
from assignments.models import Submission, Assignment


@login_required
def student_dashboard(request):

    # =========================
    # GET STUDENT PROFILE
    # =========================
    student = get_object_or_404(
        Student,
        user=request.user,
        school=request.user.school
    )

    school = student.school

    # =========================
    # CLASS DATA
    # =========================
    current_class = student.current_class
    current_stream = student.stream

    # =========================
    # SUBJECTS (FROM CLASS OR SCHOOL)
    # =========================
    subjects = Subject.objects.filter(
        school=school
    )

    # =========================
    # ASSIGNMENTS
    # =========================
    assignments = Assignment.objects.filter(
        school=school,
        class_assigned=current_class,
        is_active=True
    ).order_by("-created_at")

    # =========================
    # SUBMISSIONS
    # =========================
    submissions = Submission.objects.filter(
        student=student
    ).select_related("assignment")

    # =========================
    # ASSIGNMENT STATUS
    # =========================
    submissions_by_assignment = {s.assignment_id: s for s in submissions}
    assignments_with_status = []
    for assignment in assignments:
        submission = submissions_by_assignment.get(assignment.id)
        assignments_with_status.append({
            "assignment": assignment,
            "is_submitted": bool(submission),
            "is_graded": submission.status == "graded" if submission else False,
            "score": submission.score if submission and submission.score is not None else None,
            "feedback": submission.feedback if submission else "",
            "submission": submission,
        })

    marked_assignments = [item for item in assignments_with_status if item["is_graded"]]

    marks = StudentMark.objects.filter(
        student=student
    ).select_related("subject").order_by("-created_at")

    performance_history = get_student_term_performance_history(student)
    performance_chart = None
    subject_performance_chart = None
    if performance_history:
        labels = [entry["term_name"] for entry in performance_history]
        scores = [entry["score"] for entry in performance_history]
        performance_chart = generate_progress_chart(scores, labels=labels, title=f"{student.name} Performance Over Time", chart_type="line")

    subject_scores = []
    subject_labels = []
    for subject in subjects:
        subject_mark = StudentMark.objects.filter(student=student, subject=subject).order_by("-created_at").first()
        if subject_mark and subject_mark.marks is not None:
            subject_scores.append(float(subject_mark.marks))
            subject_labels.append(subject.short_name or subject.name[:10])

    if subject_scores:
        subject_performance_chart = generate_progress_chart(subject_scores, labels=subject_labels, title=f"{student.name} Subject Performance", chart_type="bar")

    average = 0
    if marks.exists():
        total_score = sum([float(mark.marks) for mark in marks if mark.marks is not None])
        average = round(total_score / marks.count(), 1) if marks.count() else 0

    # =========================
    # STUDENT FEE SUMMARY
    # =========================
    statement = None
    try:
        from fees.views import get_student_statement
        statement = get_student_statement(student)
    except Exception:
        statement = None

    online_classes = OnlineClass.objects.filter(
        school=school,
        class_obj=current_class
    ).filter(
        Q(stream__isnull=True) | Q(stream=current_stream)
    ).select_related("teacher", "subject", "class_obj", "stream").order_by("start_time")

    for online_class in online_classes:
        OnlineClassParticipant.objects.get_or_create(
            online_class=online_class,
            student=student,
        )

    participant_records = OnlineClassParticipant.objects.filter(
        student=student,
        online_class__in=online_classes
    ).select_related("online_class")
    participant_map = {p.online_class_id: p for p in participant_records}

    online_class_view = []
    for online_class in online_classes:
        participant = participant_map.get(online_class.id)
        online_class_view.append({
            "online_class": online_class,
            "participant": participant,
            "status": participant.status if participant else "not_tried",
        })

    # =========================
    # CONTEXT
    # =========================
    return render(request, "students/dashboard.html", {
        "student": student,
        "class": current_class,
        "stream": current_stream,
        "subjects": subjects,
        "assignments": assignments,
        "assignments_with_status": assignments_with_status,
        "marked_assignments": marked_assignments,
        "submissions": submissions,
        "marks": marks,
        "average": average,
        "performance_history": performance_history,
        "performance_chart": performance_chart,
        "subject_performance_chart": subject_performance_chart,
        "online_classes": online_class_view,
        "statement": statement,
        "notices": SchoolNotice.objects.filter(school=school).filter(recipient_type__in=['students','all']).order_by('-created_at'),
    })

def add_dorm(request):
    school = request.user.school

    if request.method == "POST":
        Dormitory.objects.create(
            name=request.POST.get("name"),
            capacity=request.POST.get("capacity"),
            supervisor=request.POST.get("supervisor"),
            school=school
        )
        messages.success(request, "Dormitory added successfully.")
        return redirect("manage_dorms")  
    return render(request, "dos/add_dorm.html")


@login_required
def student_online_class_action(request, online_class_id):
    student = get_object_or_404(
        Student,
        user=request.user,
        school=request.user.school
    )

    online_class = get_object_or_404(
        OnlineClass,
        id=online_class_id,
        school=student.school
    )

    participant, _ = OnlineClassParticipant.objects.get_or_create(
        online_class=online_class,
        student=student,
    )

    action = request.POST.get("action")
    if action == "join":
        participant.status = "joined"
        participant.joined_at = timezone.now()
        participant.save()
        messages.success(request, "You have joined the class.")
    elif action == "fail":
        participant.status = "failed"
        participant.save()
        messages.error(request, "Marked as failed to join.")

    return redirect("student_dashboard")

@login_required
def view_dorm_students(request, dorm_id):
    school = request.user.school

    dorm = get_object_or_404(Dormitory, id=dorm_id, school=school)

    students = Student.objects.filter(
        school=school,
        dormitory=dorm
    )

    return render(request, "dos/dorm_list.html", {
        "dorm": dorm,
        "students": students
    })
def edit_dorm(request, dorm_id):
        school = request.user.school

        dorm = get_object_or_404(Dormitory, id=dorm_id, school=school)

        if request.method == "POST":
            dorm.name = request.POST.get("name")
            dorm.capacity = request.POST.get("capacity")
            dorm.supervisor = request.POST.get("supervisor")
            dorm.save()

            messages.success(request, "Dormitory updated successfully.")
            return redirect("manage_dorms")

        return render(request, "dos/edit_dorm.html", {
            "dorm": dorm
        })
        
def dormitory_lists(request):
    school = request.user.school

    dorms = Dormitory.objects.filter(school=school)

    return render(request, "dos/dormitory_lists.html", {
        "dorms": dorms
    })
    
def delete_dorm(request, dorm_id):
    school = request.user.school

    dorm = get_object_or_404(Dormitory, id=dorm_id, school=school)

    if request.method == "POST":
        dorm.delete()
        messages.success(request, "Dormitory deleted successfully.")
        return redirect("manage_dorms")

    return render(request, "dos/delete_dorm.html", {
        "dorm": dorm
    })

@login_required
def print_class_list(request, class_id):
    
    class_obj = get_object_or_404(Class, id=class_id)

    students = Student.objects.filter(current_class_id=class_id)

    return render(request, "classes/print_class_list.html", {
        "class_obj": class_obj,
        "students": students
    })
import os
from io import BytesIO
from urllib.error import HTTPError, URLError
from urllib.request import urlopen
from django.conf import settings
from django.http import HttpResponse
from reportlab.platypus import PageBreak, SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph, Image
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
import datetime


def _load_reportlab_image(image_field, width, height):
    if not image_field:
        print("[logo debug] No image_field provided")
        return None

    print(f"[logo debug] Loading image_field type={type(image_field)}")

    # Helper to fetch bytes from a URL
    def _image_from_url(url):
        try:
            if not isinstance(url, str):
                return None
            url = url.strip()
            print(f"[logo debug] Trying URL fetch: {url}")
            if url.startswith("https:/") and not url.startswith("https://"):
                url = url.replace("https:/", "https://", 1)
            if url.startswith("http:/") and not url.startswith("http://"):
                url = url.replace("http:/", "http://", 1)
            parsed = urlparse(url)
            if parsed.scheme not in {"http", "https"} or not parsed.netloc:
                print(f"[logo debug] URL fetch aborted: invalid URL {url}")
                return None
            with urlopen(url) as image_file:
                image_bytes = image_file.read()
            print("[logo debug] URL fetch succeeded")
            return Image(BytesIO(image_bytes), width=width, height=height)
        except Exception as exc:
            print(f"[logo debug] URL fetch failed: {exc}")
            return None

    # 1) If image_field is a plain string it may be a URL or a Cloudinary public id
    if isinstance(image_field, str):
        print(f"[logo debug] image_field is str: {image_field}")

        # Detect a full HTTP(S) URL before treating the string as a Cloudinary public id.
        try:
            parsed = urlparse(image_field)
            if parsed.scheme in {"http", "https"} and parsed.netloc:
                img = _image_from_url(image_field)
                if img:
                    return img
                print("[logo debug] URL string fetch failed; not treating as public id")
            else:
                print("[logo debug] image_field string is not a full URL")
        except Exception as exc:
            print(f"[logo debug] URL parse failed: {exc}")

        try:
            cloud_conf = getattr(settings, "CLOUDINARY_STORAGE", None) or {}
            cloud_name = cloud_conf.get("CLOUD_NAME") if isinstance(cloud_conf, dict) else None
            if cloud_name:
                cloud_url = f"https://res.cloudinary.com/{cloud_name}/image/upload/{image_field}"
                print(f"[logo debug] Trying Cloudinary URL from public_id: {cloud_url}")
                img = _image_from_url(cloud_url)
                if img:
                    return img
        except Exception as exc:
            print(f"[logo debug] Cloudinary public id build failed: {exc}")

    image_path = None
    try:
        image_path = image_field.path
        print(f"[logo debug] image_field.path = {image_path}")
    except Exception as exc:
        print(f"[logo debug] image_field.path unavailable: {exc}")
        image_path = None

    if image_path:
        try:
            if os.path.exists(image_path):
                print("[logo debug] Loading image from local path")
                return Image(image_path, width=width, height=height)
        except Exception as exc:
            print(f"[logo debug] Local path load failed: {exc}")

    try:
        if hasattr(image_field, "file") and image_field.file:
            try:
                print("[logo debug] Loading image from file object")
                image_field.open()
                return Image(BytesIO(image_field.file.read()), width=width, height=height)
            except Exception as exc:
                print(f"[logo debug] File object load failed: {exc}")
    except Exception as exc:
        print(f"[logo debug] Error checking image_field.file: {exc}")

    public_id = None
    try:
        public_id = getattr(image_field, "public_id", None) or getattr(image_field, "publicId", None)
        print(f"[logo debug] image_field public_id = {public_id}")
    except Exception as exc:
        print(f"[logo debug] public_id lookup failed: {exc}")
        public_id = None

    if public_id:
        try:
            from cloudinary.utils import cloudinary_url
            try:
                url, _ = cloudinary_url(public_id, secure=True)
                print(f"[logo debug] Trying cloudinary_url from cloudinary.utils: {url}")
                img = _image_from_url(url)
                if img:
                    return img
            except Exception as exc:
                print(f"[logo debug] cloudinary.utils fetch failed: {exc}")
        except Exception as exc:
            print(f"[logo debug] cloudinary.utils unavailable: {exc}")
            try:
                cloud_conf = getattr(settings, "CLOUDINARY_STORAGE", None) or {}
                cloud_name = cloud_conf.get("CLOUD_NAME") if isinstance(cloud_conf, dict) else None
                if cloud_name:
                    cloud_url = f"https://res.cloudinary.com/{cloud_name}/image/upload/{public_id}"
                    print(f"[logo debug] Trying fallback Cloudinary URL: {cloud_url}")
                    img = _image_from_url(cloud_url)
                    if img:
                        return img
            except Exception as exc:
                print(f"[logo debug] Cloudinary fallback fetch failed: {exc}")

    try:
        url = getattr(image_field, "url", None)
        print(f"[logo debug] image_field.url = {url}")
    except Exception as exc:
        print(f"[logo debug] url attribute lookup failed: {exc}")
        url = None

    if url:
        img = _image_from_url(url)
        if img:
            return img

    print("[logo debug] All image load attempts failed")
    return None


@login_required
def download_class_list_pdf(request, class_id):
    """
    Export class list to PDF with proper stream filtering and professional styling.
    Ensures students are only included if they belong to the selected class/stream.
    """

    # =========================
    # GET CLASS & STREAM
    # =========================
    class_obj = get_object_or_404(Class, id=class_id)
    stream_id = request.GET.get("stream_id")
    term_id = request.GET.get("term")
    
    school = class_obj.school
    
    selected_stream = None
    if stream_id:
        selected_stream = get_object_or_404(Stream, id=stream_id, class_group=class_obj)

    # Restrict class teachers to their assigned stream for this class
    class_teacher_assignments = ClassTeacherAssignment.objects.filter(
        teacher__user=request.user,
        class_obj=class_obj
    ).select_related("stream")
    if class_teacher_assignments.exists():
        valid_stream_ids = list(class_teacher_assignments.values_list("stream_id", flat=True))
        if not selected_stream or selected_stream.id not in valid_stream_ids:
            selected_stream = Stream.objects.filter(id=valid_stream_ids[0]).first()

    # =========================
    # GET STUDENTS - STRICT FILTERING
    # =========================
    student_query = Student.objects.filter(
        current_class=class_obj,
        school=school
    )
    
    # If stream is selected, ONLY include students from that stream
    if selected_stream:
        student_query = student_query.filter(stream=selected_stream)
    else:
        # If no stream selected, exclude students who have a stream assigned
        # (only include unstreamed students)
        student_query = student_query.filter(stream__isnull=True)
    
    students = student_query.order_by("name")

    # =========================
    # BUILD FILE NAME WITH CLASS & STREAM (BETTER FORMAT)
    # =========================
    now = datetime.datetime.now()
    year = now.year
    date_str = now.strftime("%Y%m%d")
    
    # Format filename as: Grade10-East_ClassList_Term2_2026.pdf or Grade10_ClassList_General_2026.pdf
    if selected_stream and term_id:
        term_obj = Term.objects.filter(id=term_id).first()
        term_name = term_obj.name if term_obj else f"Term{term_id}"
        # Replace spaces with underscores and clean up name
        class_name_clean = class_obj.name.replace(" ", "")
        stream_name_clean = selected_stream.name.replace(" ", "")
        term_name_clean = term_name.replace(" ", "")
        filename = f"{class_name_clean}-{stream_name_clean}_ClassList_{term_name_clean}_{year}.pdf"
        display_title = f"{class_obj.name} {selected_stream.name}"
    elif selected_stream:
        class_name_clean = class_obj.name.replace(" ", "")
        stream_name_clean = selected_stream.name.replace(" ", "")
        filename = f"{class_name_clean}-{stream_name_clean}_ClassList_{year}.pdf"
        display_title = f"{class_obj.name} {selected_stream.name}"
    elif term_id:
        term_obj = Term.objects.filter(id=term_id).first()
        term_name = term_obj.name if term_obj else f"Term{term_id}"
        class_name_clean = class_obj.name.replace(" ", "")
        term_name_clean = term_name.replace(" ", "")
        filename = f"{class_name_clean}_ClassList_{term_name_clean}_{year}.pdf"
        display_title = f"{class_obj.name}"
    else:
        class_name_clean = class_obj.name.replace(" ", "")
        filename = f"{class_name_clean}_ClassList_General_{year}.pdf"
        display_title = f"{class_obj.name}"

    pagesize = landscape(A4) if term_id else A4
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagesize,
        rightMargin=25,
        leftMargin=25,
        topMargin=35,
        bottomMargin=30,
        title=display_title
    )

    styles = getSampleStyleSheet()
    elements = []

    # =========================
    # HEADER WITH LOGO & SCHOOL INFO
    # =========================
    header_data = []
    logo_image = _load_reportlab_image(getattr(school.logo, 'url', school.logo), 2.5 * cm, 2.5 * cm)

    # Build header content with school name underlined
    stream_info = f"<b>Stream:</b> {selected_stream.name}" if selected_stream else ""
    
    header_text = f"""
    <font size='18'><b>{school.name.upper()}</b></font><br/>
    <hr width="100%" noshade="1" thickness="2"/><br/>
    <font size='13'><b>OFFICIAL CLASS LIST</b></font><br/>
    <font size='11'>Class: <b>{class_obj.name}</b>&nbsp;&nbsp;&nbsp;{stream_info}</font><br/>
    <font size='10'>Academic Year: {year}</font><br/>
    <font size='9'>Generated: {now.strftime("%d %B %Y at %H:%M")}</font>
    """

    header_paragraph = Paragraph(header_text, styles["Normal"])
    
    if logo_image:
        header_table = Table(
            [[logo_image, header_paragraph]],
            colWidths=[2.5 * cm, None]
        )
    else:
        header_table = Table(
            [[header_paragraph]],
            colWidths=[None]
        )

    header_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 0.4 * cm))

    # =========================
    # STUDENT COUNT SUMMARY
    # =========================
    summary_text = f"""
    <font size='10'><b>Total Students: {students.count()}</b>&nbsp;&nbsp;&nbsp;
    Male: {students.filter(gender="Male").count()}&nbsp;&nbsp;&nbsp;
    Female: {students.filter(gender="Female").count()}</font>
    """
    summary_para = Paragraph(summary_text, styles["Normal"])
    elements.append(summary_para)
    elements.append(Spacer(1, 0.2 * cm))

    # =========================
    # BUILD STUDENT TABLE
    # =========================
    if term_id:
        # Include academic details with marks
        subjects = Subject.objects.filter(school=school).order_by("name")
        marks_qs = StudentMark.objects.filter(
            student__in=students,
            term_id=term_id
        ).select_related("subject")
        
        mark_map = {
            (mark.student_id, mark.subject_id): mark
            for mark in marks_qs
        }
        
        # Build header row
        data = [["#", "Student Name", "Admission No", "Gender"]]
        for subject in subjects:
            data[0].append(subject.short_name)
        data[0].extend(["Total", "Avg", "Grade"])
        
        # Build student rows
        for i, student in enumerate(students, start=1):
            row = [str(i), student.name, student.admission_number, student.gender]
            
            total_marks = 0
            total_subjects = 0
            
            for subject in subjects:
                mark = mark_map.get((student.id, subject.id))
                if mark:
                    marks_value = int(round(mark.marks))
                    row.append(str(marks_value))
                    total_marks += marks_value
                    total_subjects += 1
                else:
                    row.append("-")
            
            average = round(total_marks / total_subjects, 1) if total_subjects > 0 else 0
            
            # Get grade
            grade_obj = GradingPolicy.objects.filter(
                school=school,
                min_score__lte=average,
                max_score__gte=average
            ).first()
            
            grade_letter = grade_obj.short_form if grade_obj else "-"
            row.extend([str(total_marks), str(average), grade_letter])
            data.append(row)
        
        # Calculate column widths for landscape
        col_count = len(data[0])
        subject_cols = len(subjects)
        col_widths = [0.5*cm, 4*cm, 2*cm, 1.2*cm] + [0.8*cm] * subject_cols + [1*cm, 1*cm, 0.8*cm]
        
        table = Table(data, colWidths=col_widths)
    else:
        # Basic table without marks (portrait)
        data = [[
            "#",
            "Student Name",
            "Admission No",
            "Gender"
        ]]

        for i, student in enumerate(students, start=1):
            data.append([
                str(i),
                student.name,
                student.admission_number,
                student.gender
            ])

        table = Table(
            data,
            colWidths=[0.8*cm, 8*cm, 3*cm, 1.5*cm]
        )

    # Professional table styling with alternating row colors
    table_style = [
        # HEADER
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9 if term_id else 10),
        ("PADDING", (0, 0), (-1, 0), 8),
        
        # ALTERNATING ROW COLORS
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        
        # GRID
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
        
        # FONT
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8 if term_id else 9),
        
        # PADDING
        ("TOPPADDING", (0, 1), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        
        # ALIGNMENT
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (1, 0), (1, -1), "LEFT"),
    ]
    
    table.setStyle(TableStyle(table_style))
    elements.append(table)

    elements.append(Spacer(1, 0.3 * cm))

    # =========================
    # STATISTICS FOOTER
    # =========================
    total_students = students.count()
    male_students = students.filter(gender__iexact="Male").count()
    female_students = students.filter(gender__iexact="Female").count()

    stats_text = f"""
    <b>Summary:</b> Total Students: <b>{total_students}</b> 
    &nbsp;&nbsp;|&nbsp;&nbsp;
    Male: <b>{male_students}</b> 
    &nbsp;&nbsp;|&nbsp;&nbsp;
    Female: <b>{female_students}</b>
    """

    stats_para = Paragraph(stats_text, styles["Normal"])
    elements.append(stats_para)

    elements.append(Spacer(1, 0.2 * cm))

    # =========================
    # FOOTER
    # =========================
    footer_text = f"""
    <font size='8'><i>Generated by Brainet ERP System | {now.strftime("%d-%m-%Y %H:%M")} | Classification: Official Record</i></font>
    """
    footer_para = Paragraph(footer_text, styles["Normal"])
    elements.append(footer_para)

    # =========================
    # BUILD PDF
    # =========================
    doc.build(elements)
    buffer.seek(0)

    # =========================
    # RETURN DOWNLOAD RESPONSE
    # =========================
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/pdf"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
from django.shortcuts import get_object_or_404, render
from students.models import Student

def student_report(request, student_id):
    student = get_object_or_404(Student, id=student_id)

    # later we will attach marks, exams, etc.
    return render(request, "students/report.html", {
        "student": student
    })
    
def class_lists(request, class_id):
    school = request.user.school

    school_class = get_object_or_404(
        Class,
        id=class_id,
        school=school
    )

    students = Student.objects.filter(
        school=school,
        current_class=school_class
    ) 

    return render(request, "dos/class_list.html", {
        "school_class": school_class,
        "students": students
    })

@login_required
def class_list_preview(request, class_id):
    """Preview class list with academic details (marks, grades, points)"""
    school = request.user.school
    term_id = request.GET.get("term")
    
    school_class = get_object_or_404(Class, id=class_id, school=school)
    
    students = Student.objects.filter(
        school=school,
        current_class=school_class,
        status="active"
    ).select_related("stream").order_by("name")
    
    subjects = Subject.objects.filter(school=school).order_by("name")
    
    preview_rows = []
    
    if term_id:
        # Get all marks for this class and term
        marks_qs = StudentMark.objects.filter(
            student__in=students,
            term_id=term_id
        ).select_related("subject")
        
        mark_map = {
            (mark.student_id, mark.subject_id): mark
            for mark in marks_qs
        }
        
        for student in students:
            subject_scores = []
            total_marks = 0
            total_subjects = 0
            total_points = 0
            
            for subject in subjects:
                mark = mark_map.get((student.id, subject.id))
                
                if mark is not None:
                    marks_value = int(round(mark.marks))
                    total_marks += marks_value
                    total_subjects += 1
                    if mark.points:
                        total_points += mark.points
                else:
                    marks_value = None
                
                subject_scores.append(marks_value)
            
            average = round(total_marks / total_subjects, 2) if total_subjects > 0 else 0
            
            # Get grade
            grade_obj = GradingPolicy.objects.filter(
                school=school,
                min_score__lte=average,
                max_score__gte=average
            ).first()
            
            preview_rows.append({
                "student": student,
                "stream_name": student.stream.name if student.stream else "",
                "subject_scores": subject_scores,
                "total": total_marks,
                "average": average,
                "grade": grade_obj.short_form if grade_obj else "-",
                "points": total_points,
            })
        
        # Sort by total descending
        preview_rows.sort(key=lambda x: x["total"], reverse=True)
        
        # Add rank
        for idx, row in enumerate(preview_rows, start=1):
            row["rank"] = idx
    
    # Get terms for dropdown
    terms = Term.objects.filter(school=school).order_by("-id")
    selected_term = None
    if term_id:
        selected_term = Term.objects.filter(id=term_id).first()
    
    return render(request, "dos/class_list_preview.html", {
        "school_class": school_class,
        "students": students,
        "subjects": subjects,
        "preview_rows": preview_rows,
        "terms": terms,
        "selected_term": selected_term,
        "term_id": term_id,
    })
    

def view_school(request, school_id):
    school = get_object_or_404(School, id=school_id)

    return render(request, "dos/view_school.html", {
        "school": school
    })  
@login_required
def create_staff(request):
    if request.method == "POST":
        name = request.POST.get("name")
        email = request.POST.get("email")
        phone = request.POST.get("phone")
        role = request.POST.get("role")
        school_id = request.POST.get("school_id")

        school = get_object_or_404(School, id=school_id)

        # AUTO USERNAME
        username = email.split("@")[0]

        # Use your CustomUser model
        user = CustomUser.objects.create_user(
            username=username,
            email=email,
            password="password123",   # or generate random
            role=role,
            school=school
        )

        # Save extra fields
        user.first_name = name
        user.save()

        send_user_verification_email(user, request=request, role_name=role.replace('_', ' ').title())

        messages.success(
            request,
            f"{role.replace('_', ' ').title()} account created successfully. A verification email has been sent."
        )

        return redirect("superuser_dashboard")

    # If GET, render form
    schools = School.objects.all()
    return render(request, "schools/create_staff.html", {"schools": schools})


def manage_staff(request):
    school = request.user.school

    staff = CustomUser.objects.filter(school=school).exclude(is_superuser=True)

    return render(request, "dos/staff.html", {
        "staff": staff
    })


@login_required
def create_bursar(request):
    """Principal creates and manages bursar accounts for their school."""
    if getattr(request.user, 'role', None) != 'principal' or not request.user.school:
        messages.error(request, 'Only principals can create bursar accounts.')
        return redirect('dashboard')

    school = request.user.school
    selected_bursar = None
    bursars = CustomUser.objects.filter(school=school, role='bursar').order_by('first_name', 'email')

    if request.method == 'POST':
        action = request.POST.get('action', 'create')
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        password = request.POST.get('password', '')

        if action == 'update':
            bursar_id = request.POST.get('bursar_id')
            if not bursar_id:
                messages.error(request, 'Invalid bursar selected.')
                return redirect('create_bursar')

            bursar = get_object_or_404(CustomUser, id=bursar_id, school=school, role='bursar')

            if not all([name, email, phone]):
                messages.error(request, 'All fields are required.')
                return redirect(f'{reverse("create_bursar")}?bursar_id={bursar.id}')

            if CustomUser.objects.exclude(id=bursar.id).filter(email=email).exists():
                messages.error(request, f'Email {email} is already registered.')
                return redirect(f'{reverse("create_bursar")}?bursar_id={bursar.id}')

            bursar.first_name = name
            bursar.email = email
            bursar.last_name = ''
            if hasattr(bursar, 'phone'):
                bursar.phone = phone
            if password:
                bursar.set_password(password)
            bursar.save()

            messages.success(request, f'Bursar account updated for {name}.')
            return redirect('create_bursar')

        # CREATE
        if not all([name, email, phone]):
            messages.error(request, 'All fields are required.')
            return redirect('create_bursar')

        if CustomUser.objects.filter(email=email).exists():
            messages.error(request, f'Email {email} is already registered.')
            return redirect('create_bursar')

        try:
            first_name = name
            last_name = ''
            if ' ' in name:
                parts = name.split()
                first_name = parts[0]
                last_name = ' '.join(parts[1:])

            user = CustomUser.objects.create_user(
                email=email,
                password='bursar123',
                role='bursar',
                school=school,
                first_name=first_name,
                last_name=last_name,
                email_verified=False,
            )
            if hasattr(user, 'phone'):
                user.phone = phone
            user.save()

            send_user_verification_email(user, request=request, role_name='Bursar')
            messages.success(request, f'Bursar account created for {name}. A verification email has been sent to {email}.')
            return redirect('create_bursar')
        except Exception as e:
            messages.error(request, f'Error creating bursar account: {str(e)}')
            return redirect('create_bursar')

    bursar_id = request.GET.get('bursar_id')
    if bursar_id:
        selected_bursar = get_object_or_404(CustomUser, id=bursar_id, school=school, role='bursar')

    return render(request, 'schools/create_bursar.html', {
        'bursars': bursars,
        'selected_bursar': selected_bursar,
    })


@login_required
def delete_bursar(request, user_id):
    if getattr(request.user, 'role', None) != 'principal' or not request.user.school:
        messages.error(request, 'Only principals can delete bursar accounts.')
        return redirect('dashboard')

    school = request.user.school
    bursar = get_object_or_404(CustomUser, id=user_id, school=school, role='bursar')

    if request.method == 'POST':
        name = bursar.get_full_name() or bursar.email
        bursar.delete()
        messages.success(request, f'{name} deleted successfully.')

    return redirect('create_bursar')

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required

from users.models import CustomUser
from .models import School, DirectorOfStudies, Principal


@login_required
def register_dos_by_superuser(request):

    if request.method == "POST":

        try:

            name = request.POST.get("name")
            email = request.POST.get("email")
            phone = request.POST.get("phone")
            password = request.POST.get("password")
            school_id = request.POST.get("school")

            # VALIDATION
            if not all([name, email, phone, password, school_id]):

                messages.error(
                    request,
                    "Please fill all required fields."
                )

                return redirect("superuser_dashboard")

            school = School.objects.get(id=school_id)

            # EMAIL EXISTS
            if CustomUser.objects.filter(email=email).exists():

                messages.warning(
                    request,
                    "Email already taken."
                )

                return redirect("superuser_dashboard")

            # PHONE EXISTS
            if DirectorOfStudies.objects.filter(phone=phone).exists():

                messages.warning(
                    request,
                    "Phone number already used."
                )

                return redirect("superuser_dashboard")

            # SCHOOL ALREADY HAS DOS
            if DirectorOfStudies.objects.filter(school=school).exists():

                messages.warning(
                    request,
                    "This school already has a DOS account."
                )

                return redirect("superuser_dashboard")

            # CREATE USER (initially unverified)
            user = User.objects.create_user(
            email=email,
            password=password,
            role="dos",
            school=school,
            email_verified=False  # Must verify via code before full access

            )

            # CREATE DOS PROFILE
            DirectorOfStudies.objects.create(
                user=user,
                school=school,
                name=name,
                email=email,
                phone=phone
            )

            send_user_verification_email(user, request=request, role_name='Director of Studies')

            messages.success(
                request,
                f"{name} registered successfully. A verification email has been sent."
            )

        except School.DoesNotExist:

            messages.error(
                request,
                "Selected school does not exist."
            )

        except IntegrityError:

            messages.error(
                request,
                "Duplicate information detected."
            )

        except Exception as e:

            messages.error(
                request,
                f"System Error: {str(e)}"
            )

    return redirect("superuser_dashboard")
@login_required
def register_principal_by_superuser(request):
    if request.method == "POST":
        try:
            name = request.POST.get("name")
            email = request.POST.get("email")
            phone = request.POST.get("phone")
            password = request.POST.get("password")
            school_id = request.POST.get("school")

            # VALIDATION
            if not all([name, email, phone, password, school_id]):
                messages.error(request, "Please fill all required fields.")
                return redirect("superuser_dashboard")

            school = School.objects.get(id=school_id)

            # EMAIL EXISTS
            if CustomUser.objects.filter(email=email).exists():
                messages.warning(request, "Email already taken.")
                return redirect("superuser_dashboard")

            # PHONE EXISTS
            if Principal.objects.filter(phone=phone).exists():
                messages.warning(request, "Phone number already used.")
                return redirect("superuser_dashboard")

            # SCHOOL ALREADY HAS PRINCIPAL
            if Principal.objects.filter(school=school).exists():
                messages.warning(request, "This school already has a Principal account.")
                return redirect("superuser_dashboard")

            # CREATE USER (initially unverified)
            user = CustomUser.objects.create_user(
                email=email,
                password=password,
                role="principal",
                school=school,
                email_verified=False  # Must verify via code before full access
            )

            # CREATE PRINCIPAL PROFILE
            Principal.objects.create(
                user=user,
                school=school,
                name=name,
                email=email,
                phone=phone
            )

            send_user_verification_email(user, request=request, role_name='Principal')

            messages.success(request, f"{name} registered successfully as Principal. A verification email has been sent.")

        except School.DoesNotExist:
            messages.error(request, "Selected school does not exist.")

        except IntegrityError:
            messages.error(request, "Duplicate information detected.")

        except Exception as e:
            messages.error(request, f"System Error: {str(e)}")

    return redirect("superuser_dashboard")

def activate_school(request, school_id):
    school = get_object_or_404(School, id=school_id)
    school.is_active = True
    school.is_verified = True
    school.verified_at = timezone.now()
    if hasattr(request.user, 'email'):
        school.verified_by = request.user
    school.save()

    email_subject = f"School verified: {school.name}"
    email_body = (
        f"The school '{school.name}' has been verified and activated by {request.user.get_full_name() or request.user.email}.\n"
        f"School email: {school.email}\n"
        f"Phone: {school.phone}\n"
    )

    try:
        if settings.EMAIL_HOST_USER and school.email:
            send_email(
                to_email=[school.email],
                subject=email_subject,
                message=email_body,
                recipient_name=school.name,
                html=False,
            )
        for su in User.objects.filter(is_superuser=True):
            if su.email:
                send_email(
                    to_email=[su.email],
                    subject=email_subject,
                    message=email_body,
                    recipient_name=su.get_full_name() or su.email,
                    html=False,
                )
    except Exception:
        pass

    messages.success(request, "School activated and verified successfully.")
    return redirect("view_school", school_id=school.id)


def verify_school_via_token(request, token):
    school = get_object_or_404(School, verification_token=token)

    if not school.verification_sent_at or school.verification_sent_at + timedelta(hours=1) < timezone.now():
        school.verification_token = None
        school.save(update_fields=["verification_token"])
        messages.error(request, "School verification link has expired. Please request a new school registration email.")
        return redirect("register_school")

    school.is_verified = True
    school.verified_at = timezone.now()
    school.verification_token = None
    school.save(update_fields=["is_verified", "verified_at", "verification_token"])
    messages.success(request, "School email verified successfully. A school administrator will complete activation soon.")
    return redirect("register_school_success")

def active_schools(request):
    schools = School.objects.filter(is_active=True)
    return render(request, "dos/active_schools.html", {
        "schools": schools
    })
    
@login_required
def deactivate_school(request, school_id):
    """Deactivate a school and redirect to deactivation page"""
    school = get_object_or_404(School, id=school_id)
    
    if request.method == "POST":
        reason = request.POST.get("reason", "")
        school.is_active = False
        school.license_status = 'suspended'
        school.deactivated_at = timezone.now()
        school.deactivation_reason = reason
        school.save()
        messages.success(request, "School has been deactivated.")
        return redirect("school_deactivated", school_id=school.id)
    
    context = {'school': school}
    return render(request, "schools/deactivate_confirm.html", context)

@login_required
def school_deactivated(request, school_id):
    """Page shown when school is deactivated"""
    school = get_object_or_404(School, id=school_id)
    
    # Check if user has permission to view this page (superuser, DOS, or Principal)
    if not (
        request.user.is_superuser or 
        (hasattr(request.user, 'dos_profile') and request.user.dos_profile.school_id == school_id) or
        (hasattr(request.user, 'principal') and request.user.principal.school_id == school_id)
    ):
        return redirect("landing_page")
    
    pending_renewals = school.license_renewals.filter(status='pending').select_related('requested_by')
    context = {
        'school': school,
        'has_pending_renewal': pending_renewals.exists(),
        'pending_renewals': pending_renewals,
    }
    return render(request, "schools/school_deactivated.html", context)

@login_required
def request_license_renewal(request, school_id):
    """Request license renewal"""
    from .models import LicenseRenewal
    
    school = get_object_or_404(School, id=school_id)
    
    # Check permission
    if not (request.user.is_superuser or 
            (hasattr(request.user, 'dos_profile') and request.user.dos_profile.school_id == school_id) or
            (hasattr(request.user, 'principal') and request.user.principal.school_id == school_id)):
        return redirect("landing_page")
    
    if request.method == "POST":
        renewal_period = request.POST.get("renewal_period", "365")
        
        LicenseRenewal.objects.create(
            school=school,
            requested_by=request.user,
            renewal_period_days=int(renewal_period)
        )
        # Notify DOS and superusers about the renewal request
        try:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            superusers = User.objects.filter(is_superuser=True)

            title = "License renewal requested"
            message_text = f"{request.user.get_full_name() or request.user.email} requested license renewal for {school.name}."

            # notify each superuser
            for su in superusers:
                Notification.objects.create(
                    school=school,
                    sender=request.user,
                    recipient=su,
                    title=title,
                    message=message_text
                )

            # notify school's DOS (if assigned)
            dos = getattr(school, 'dos', None)
            if dos and getattr(dos, 'user', None):
                Notification.objects.create(
                    school=school,
                    sender=request.user,
                    recipient=dos.user,
                    title=title,
                    message=message_text
                )
        except Exception:
            pass
        
        messages.success(request, "License renewal request submitted. Please wait for approval.")
        return redirect("school_deactivated", school_id=school.id)
    
    context = {'school': school}
    return render(request, "schools/request_renewal.html", context)

@login_required
def approve_license_renewal(request, renewal_id):
    """Superuser approval for license renewal"""
    from .models import LicenseRenewal
    
    if not request.user.is_superuser:
        return redirect("landing_page")
    
    renewal = get_object_or_404(LicenseRenewal, id=renewal_id)
    
    if request.method == "POST":
        approval_status = request.POST.get("status")
        notes = request.POST.get("notes", "")
        
        renewal.status = approval_status
        renewal.processed_by = request.user
        renewal.processed_at = timezone.now()
        renewal.notes = notes
        renewal.save()
        
        if approval_status == 'approved':
            # Activate school and extend license
            school = renewal.school
            school.is_active = True
            school.license_status = 'active'
            school.license_expiry = timezone.now().date() + timedelta(days=renewal.renewal_period_days)
            school.deactivated_at = None
            school.save()
            
            messages.success(request, f"License renewed for {school.name} until {school.license_expiry}")
        else:
            messages.info(request, "License renewal has been rejected.")
        
        return redirect("superuser_dashboard")
    
    context = {'renewal': renewal}
    return render(request, "schools/approve_renewal.html", context)

@superuser_required
def add_school(request):
    if request.method == "POST":
        name = request.POST.get("name")
        address = request.POST.get("address")
        phone = request.POST.get("phone")
        email = request.POST.get("email")
        subscription_balance = request.POST.get("subscription_balance") or 0
        bank_name = request.POST.get("bank_name") or None
        account_number = request.POST.get("account_number") or None
        logo = request.FILES.get("logo")

        logger = logging.getLogger(__name__)
        try:
            school = School.objects.create(
                name=name,
                address=address,
                phone=phone,
                email=email,
                subscription_balance=subscription_balance,
                bank_name=bank_name,
                account_number=account_number,
                is_active=False,
            )
            if logo:
                school.logo = logo

            messages.success(request, "School added successfully.")
            messages.warning(
                request,
                "New school accounts expire within 48 hours if not activated. Contact admin to activate."
            )
        except (OSError, IOError) as e:
            # Likely running on a read-only filesystem (serverless). Create record without logo
            logger.exception("Failed to save uploaded logo file for school: %s", e)
            school = School.objects.create(
                name=name,
                address=address,
                phone=phone,
                email=email,
                subscription_balance=subscription_balance,
                bank_name=bank_name,
                account_number=account_number,
                is_active=False,
            )
            messages.success(request, "School added, but logo upload failed.")
            messages.error(request, "Logo could not be saved on this server. Configure remote media storage (S3/GCS) and try again.")

    return redirect("superuser_dashboard")


def register_school(request):
    if request.method == "POST":
        name = request.POST.get("name")
        address = request.POST.get("address")
        phone = request.POST.get("phone")
        email = request.POST.get("email")
        admin_name = request.POST.get("admin_name")
        admin_email = request.POST.get("admin_email")
        admin_phone = request.POST.get("admin_phone")
        admin_password = request.POST.get("admin_password") or request.POST.get("password")

        if not all([name, address, phone, email, admin_name, admin_email, admin_phone, admin_password]):
            messages.error(request, "Please fill in all required school and admin details.")
            return render(request, "schools/register_school.html", {
                "name": name,
                "address": address,
                "phone": phone,
                "email": email,
                "admin_name": admin_name,
                "admin_email": admin_email,
                "admin_phone": admin_phone,
            })

        if School.objects.filter(email=email).exists():
            messages.error(request, "Email already taken.")
            return render(request, "schools/register_school.html", {
                "name": name,
                "address": address,
                "phone": phone,
                "email": email,
                "admin_name": admin_name,
                "admin_email": admin_email,
                "admin_phone": admin_phone,
            })

        if CustomUser.objects.filter(email=admin_email).exists():
            return render(request, "schools/register_school.html", {
                "name": name,
                "address": address,
                "phone": phone,
                "email": email,
                "admin_name": admin_name,
                "admin_email": admin_email,
                "admin_phone": admin_phone,
            })

        if Principal.objects.filter(phone=admin_phone).exists():
            messages.error(request, "This admin phone number is already in use.")
            return render(request, "schools/register_school.html", {
                "name": name,
                "address": address,
                "phone": phone,
                "email": email,
                "admin_name": admin_name,
                "admin_email": admin_email,
                "admin_phone": admin_phone,
            })

        school = School.objects.create(
            name=name,
            address=address,
            phone=phone,
            email=email,
            is_active=True,
            is_verified=True,
            license_status='active',
            license_expiry=timezone.now().date() + timedelta(days=5),
        )

        user = CustomUser.objects.create_user(
            email=admin_email,
            password=admin_password,
            role="principal",
            school=school,
            email_verified=True,
        )
        Principal.objects.create(
            user=user,
            school=school,
            name=admin_name,
            email=admin_email,
            phone=admin_phone,
        )

        send_school_verification_email(school, request=request)

        login_link = request.build_absolute_uri(reverse("login")) if request else reverse("login")
        admin_message = (
            f"Hello {admin_name},\n\n"
            f"Your Brainet school admin account for {school.name} has been created.\n"
            f"Email: {admin_email}\n"
            f"Temporary password: {admin_password}\n\n"
            f"Use this link to sign in: {login_link}\n\n"
            "Please change your password after your first login."
        )
        send_email(
            to_email=admin_email,
            subject="Your Brainet school admin account",
            message=admin_message,
            recipient_name=admin_name,
            html=False,
        )

        try:
            superusers = User.objects.filter(is_superuser=True)
            sender = superusers.first() if superusers.exists() else None
            title = "New school registration request"
            message_text = (
                f"A new school registration request has been submitted for '{school.name}'. "
                f"School admin: {admin_name} ({admin_email})."
            )
            for su in superusers:
                Notification.objects.create(
                    school=school,
                    sender=sender or su,
                    recipient=su,
                    title=title,
                    message=message_text,
                )
                if settings.EMAIL_HOST_USER and su.email:
                    try:
                        send_email(
                            to_email=[su.email],
                            subject=title,
                            message=message_text,
                            recipient_name=su.get_full_name() or su.email,
                            html=False,
                        )
                    except Exception:
                        pass
        except Exception:
            pass

        messages.success(
            request,
            f"Thank you. Your school trial account has been created for {school.name}. The free trial ends after 5 days on {school.license_expiry}. The school admin login details were sent to the admin email."
        )
        return redirect("register_school_success")

    return render(request, "schools/register_school.html")


def register_school_success(request):
    return render(request, "schools/register_school_success.html")

def manage_schools(request):
    schools = School.objects.all()

    return render(request, "schools/schools.html", {
        "schools": schools
    })
    
    
def view_school(request, school_id):
    school = get_object_or_404(School, id=school_id)

    return render(request, "schools/view_school.html", {
        "school": school
    })
    
def edit_school(request, school_id):
    school = get_object_or_404(School, id=school_id)

    if request.method == "POST":
        school.name = request.POST.get("name")
        school.address = request.POST.get("address")
        school.phone = request.POST.get("phone")
        school.email = request.POST.get("email")
        school.save()

        messages.success(request, "School details updated successfully.")
        return redirect("view_school", school_id=school.id)

    return render(request, "schools/edit_school.html", {
        "school": school
    })
    
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages

from .models import School


def delete_school(request, school_id):

    school = get_object_or_404(
        School,
        id=school_id
    )

    try:
        school.delete()

        messages.success(
            request,
            "School deleted successfully."
        )

    except Exception as e:

        messages.error(
            request,
            f"Delete failed: {str(e)}"
        )

    return redirect("superuser_dashboard")
    
def manage_terms(request):
    school = request.user.school

    terms = Term.objects.filter(school=school)

    return render(request, "dos/terms.html", {
        "terms": terms
    })  

@login_required
@exam_controller_required
@require_POST
def open_exam_window(request):
    school = request.user.school
    exams = Exam.objects.filter(school=school)
    active_count = exams.filter(is_active=True).count()

    if active_count > 0:
        messages.info(request, "Exam window is already open.")
        set_exam_window_state(request, school, True)
    else:
        exams.update(is_active=True)
        set_exam_window_state(request, school, True)
        messages.success(request, "Exam window opened successfully. Marks entry is now available.")

    next_url = request.POST.get("next") or request.GET.get("next")
    if next_url:
        return redirect(next_url)
    return redirect("principal_dashboard" if getattr(request.user, "role", None) == "principal" else "dos_dashboard")


@login_required
@exam_controller_required
@require_POST
def close_exam_window(request):
    school = request.user.school
    exams = Exam.objects.filter(school=school)
    active_count = exams.filter(is_active=True).count()

    if active_count == 0:
        messages.info(request, "Exam window is already closed.")
        set_exam_window_state(request, school, False)
    else:
        exams.update(is_active=False)
        set_exam_window_state(request, school, False)
        messages.warning(request, "Exam window closed. Contact admin to enter or update results if this was a mistake.")

    next_url = request.POST.get("next") or request.GET.get("next")
    if next_url:
        return redirect(next_url)
    return redirect("principal_dashboard" if getattr(request.user, "role", None) == "principal" else "dos_dashboard")
from django.shortcuts import render, redirect
from django.contrib import messages

from exams.models import Exam, ExamSubject, Mark
from schools.models import Subject, Term
from django.contrib import messages
from django.shortcuts import redirect, render

@login_required
def manage_exams(request):

    school = request.user.school

    terms = Term.objects.filter(
        school=school
    )

    subjects = Subject.objects.filter(
        school=school
    )

    exams = Exam.objects.filter(
        school=school
    ).order_by("-created_at")

    edit_exam = None
    edit_subject_ids = []
    edit_exam_id = request.GET.get("edit_exam_id")
    delete_exam_id = request.GET.get("delete_exam_id")

    if delete_exam_id:
        exam_to_delete = Exam.objects.filter(id=delete_exam_id, school=school).first()
        if exam_to_delete:
            exam_to_delete.delete()
            messages.success(request, "Exam deleted successfully.")
        return redirect("manage_exams")

    if edit_exam_id:
        edit_exam = Exam.objects.filter(id=edit_exam_id, school=school).first()
        if edit_exam:
            edit_subject_ids = list(edit_exam.exam_subjects.values_list("subject_id", flat=True))

    if request.method == "POST":

        try:

            exam_id = request.POST.get("exam_id")
            name = request.POST.get("name")
            term_id = request.POST.get("term")
            exam_type = request.POST.get("exam_type")

            subject_ids = request.POST.getlist(
                "subjects"
            )

            # Require at least one subject when creating/updating an exam
            if not subject_ids:
                messages.error(request, "Please select at least one subject for the exam.")
                return redirect("manage_exams")

            if exam_id:
                exam = Exam.objects.filter(id=exam_id, school=school).first()
                if exam:
                    exam.name = name
                    exam.term_id = term_id
                    exam.exam_type = exam_type
                    exam.save()
                    ExamSubject.objects.filter(exam=exam).delete()
                    message_text = "Exam updated successfully."
                else:
                    exam = Exam.objects.create(
                        school=school,
                        name=name,
                        term_id=term_id,
                        exam_type=exam_type
                    )
                    message_text = "Exam created successfully."
            else:
                exam = Exam.objects.create(
                    school=school,
                    name=name,
                    term_id=term_id,
                    exam_type=exam_type
                )
                message_text = "Exam created successfully."

            for subject_id in subject_ids:
                ExamSubject.objects.create(
                    exam=exam,
                    subject_id=subject_id
                )

            messages.success(
                request,
                message_text
            )

            return redirect("manage_exams")

            messages.success(
                request,
                "Exam created successfully."
            )

            return redirect("manage_exams")

        except Exception as e:

            messages.error(
                request,
                f"System Error: {str(e)}"
            )

    return render(
        request,
        "dos/manage_exams.html",
        {
            "terms": terms,
            "subjects": subjects,
            "exams": exams,
        }
    )
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction

@login_required
def enter_marks(request):

    school = request.user.school

    # =========================
    # LOAD DATA (SCHOOL SAFE)
    # =========================
    classes = Class.objects.filter(school=school).order_by("name")
    terms = Term.objects.filter(school=school).order_by("name")
    exams = Exam.objects.filter(school=school).order_by("name")

    # =========================
    # GET FILTERS
    # =========================
    selected_class = request.GET.get("class")
    selected_option = request.GET.get("option", "all")
    selected_stream = request.GET.get("stream")
    selected_subject = request.GET.get("subject")
    selected_exam = request.GET.get("exam")
    selected_term = request.GET.get("term")
    selected_entry_mode = request.GET.get("entry_mode", "single")

    students = Student.objects.none()
    streams = Stream.objects.none()
    subjects = Subject.objects.filter(school=school).order_by("name")
    selected_exam_obj = None
    existing_marks = {}
    exam_closed_warning = None
    exam_window_open, _ = resolve_exam_window_state(request, school)

    # =========================
    # LOAD STREAMS AND STUDENTS SAFELY
    # =========================
    if selected_class:
        streams = Stream.objects.filter(class_group_id=selected_class).order_by("name")
        students_qs = Student.objects.filter(
            school=school,
            current_class_id=selected_class
        )

        if selected_stream:
            students_qs = students_qs.filter(stream_id=selected_stream)

        students = list(students_qs.order_by("name"))

        # Filter subjects by class + option selection
        class_subjects = ClassSubject.objects.filter(class_name_id=selected_class)

        if selected_option == "optional":
            class_subjects = class_subjects.filter(is_optional=True)
        elif selected_option == "core":
            class_subjects = class_subjects.filter(is_optional=False)

        subjects = Subject.objects.filter(
            id__in=class_subjects.values_list("subject_id", flat=True)
        ).order_by("name")

    # If an exam is selected, restrict subjects to those registered for the exam
    if selected_exam:
        try:
            selected_exam_obj = Exam.objects.filter(id=selected_exam, school=school).first()
            if selected_exam_obj:
                exam_subject_ids = selected_exam_obj.exam_subjects.values_list("subject_id", flat=True)
                subjects = subjects.filter(id__in=exam_subject_ids).order_by("name")
                if not subjects.exists():
                    messages.warning(request, "Selected exam has no registered subjects. Please assign subjects to the exam before entering marks.")
                if not selected_exam_obj.is_active:
                    exam_closed_warning = "This exam is closed. Contact admin to enter or update results."
                    messages.warning(request, exam_closed_warning)
        except Exception:
            pass

    if not exam_window_open and not selected_exam_obj:
        exam_closed_warning = "Exam entry is currently closed. Contact your school admin to open the exam window."
        messages.warning(request, exam_closed_warning)

    if selected_class and selected_term and students:
        term = Term.objects.filter(id=selected_term, school=school).first()
        if term:
            if selected_entry_mode == "bulk":
                marks_qs = StudentMark.objects.filter(
                    student__in=[student.id for student in students],
                    term=term,
                )
                if selected_exam_obj:
                    marks_qs = marks_qs.filter(exam=selected_exam_obj)
                marks_qs = marks_qs.order_by("student_id", "subject_id", "-created_at")

                latest_marks = {}
                for mark_obj in marks_qs:
                    student_id = getattr(mark_obj.student_id, "__int__", lambda: mark_obj.student_id)()
                    subject_id = getattr(mark_obj.subject_id, "__int__", lambda: mark_obj.subject_id)()
                    key = (student_id, subject_id)
                    if key in latest_marks:
                        continue
                    latest_marks[key] = str(mark_obj.marks) if mark_obj and mark_obj.marks is not None else ""

                for student in students:
                    subject_values = []
                    for subject in subjects:
                        current_value = latest_marks.get((student.id, subject.id), "")
                        subject_values.append({"subject": subject, "current_mark": current_value})
                    setattr(student, "subject_entry_values", subject_values)
            elif selected_subject:
                subject = Subject.objects.filter(id=selected_subject, school=school).first()
                if subject:
                    marks_qs = StudentMark.objects.filter(
                        student__in=[student.id for student in students],
                        subject=subject,
                        term=term,
                    )
                    if selected_exam_obj:
                        marks_qs = marks_qs.filter(exam=selected_exam_obj)
                    marks_qs = marks_qs.order_by("student_id", "-created_at")

                    seen_students = set()
                    for mark_obj in marks_qs:
                        student_id = getattr(mark_obj.student_id, "__int__", lambda: mark_obj.student_id)()
                        if student_id in seen_students:
                            continue
                        seen_students.add(student_id)
                        existing_marks[student_id] = str(mark_obj.marks) if mark_obj and mark_obj.marks is not None else ""

                    for student in students:
                        setattr(student, "current_mark", existing_marks.get(getattr(student, "id", None), ""))

    # =========================
    # SAVE MARKS
    # =========================
    if request.method == "POST":

        class_id = request.POST.get("class")
        option = request.POST.get("option", "all")
        stream_id = request.POST.get("stream")
        subject_id = request.POST.get("subject")
        exam_id = request.POST.get("exam")
        term_id = request.POST.get("term")
        entry_mode = request.POST.get("entry_mode", "single")

        # -------------------------
        # VALIDATION
        # -------------------------
        if not class_id or not term_id:
            messages.error(request, "Class and Term are required")
            return redirect("enter_marks")

        term = Term.objects.filter(id=term_id, school=school).first()
        exam = Exam.objects.filter(id=exam_id, school=school).first() if exam_id else None

        if not term:
            messages.error(request, "Invalid term selected")
            return redirect("enter_marks")

        if not exam_id:
            messages.error(request, "Please select an active exam before saving marks.")
            redirect_url = f"/schools/principal/enter-marks/?class={class_id}&term={term_id}"
            if stream_id:
                redirect_url += f"&stream={stream_id}"
            if option and option != "all":
                redirect_url += f"&option={option}"
            if entry_mode:
                redirect_url += f"&entry_mode={entry_mode}"
            return redirect(redirect_url)

        if not exam:
            messages.error(request, "Selected exam was not found. Please choose a valid exam.")
            redirect_url = f"/schools/principal/enter-marks/?class={class_id}&term={term_id}"
            if stream_id:
                redirect_url += f"&stream={stream_id}"
            if exam_id:
                redirect_url += f"&exam={exam_id}"
            if option and option != "all":
                redirect_url += f"&option={option}"
            if entry_mode:
                redirect_url += f"&entry_mode={entry_mode}"
            return redirect(redirect_url)

        if not exam.is_active:
            messages.error(request, "This exam is closed. Contact admin to enter or update results.")
            redirect_url = f"/schools/principal/enter-marks/?class={class_id}&term={term_id}"
            if stream_id:
                redirect_url += f"&stream={stream_id}"
            if exam_id:
                redirect_url += f"&exam={exam_id}"
            if option and option != "all":
                redirect_url += f"&option={option}"
            if entry_mode:
                redirect_url += f"&entry_mode={entry_mode}"
            return redirect(redirect_url)

        grading = GradingPolicy.objects.filter(school=school)
        students_for_entry = Student.objects.filter(
            school=school,
            current_class_id=class_id
        )
        if stream_id:
            students_for_entry = students_for_entry.filter(stream_id=stream_id)
        students_for_entry = students_for_entry.order_by("name")

        if entry_mode == "bulk":
            class_subjects = ClassSubject.objects.filter(class_name_id=class_id)
            if option == "optional":
                class_subjects = class_subjects.filter(is_optional=True)
            elif option == "core":
                class_subjects = class_subjects.filter(is_optional=False)

            subject_ids_for_entry = list(class_subjects.values_list("subject_id", flat=True))
            if exam:
                subject_ids_for_entry = list(
                    Subject.objects.filter(id__in=exam.exam_subjects.values_list("subject_id", flat=True), school=school)
                    .values_list("id", flat=True)
                )
            if not subject_ids_for_entry:
                messages.error(request, "No subjects are assigned for the selected class or exam")
                return redirect("enter_marks")
        else:
            if not subject_id:
                messages.error(request, "Subject is required for single-entry mode")
                return redirect("enter_marks")
            subject = Subject.objects.filter(id=subject_id, school=school).first()
            if not subject:
                messages.error(request, "Invalid subject selected")
                return redirect("enter_marks")
            subject_ids_for_entry = [subject.id]

        saved_count = 0

        for key, value in request.POST.items():
            if not key.startswith("mark_"):
                continue

            match = re.fullmatch(r"mark_(\d+)(?:_(\d+))?", key)
            if not match:
                continue

            student_id = int(match.group(1))
            student = Student.objects.filter(id=student_id, school=school).first()
            if not student:
                logger.debug("Skipping mark for unknown student id=%s", student_id)
                continue

            raw = (value or "").strip()
            if raw == "":
                logger.debug("Empty mark input for student id=%s, skipping", student_id)
                continue

            try:
                marks = int(raw)
            except (ValueError, TypeError):
                logger.debug("Invalid non-integer mark for student id=%s: %s", student_id, raw)
                continue

            if marks < 0 or marks > 100:
                logger.debug("Out of range mark for student id=%s: %s", student_id, marks)
                continue

            parsed_subject_id = match.group(2)
            if entry_mode == "bulk":
                if not parsed_subject_id:
                    continue
                subject_for_entry = Subject.objects.filter(id=parsed_subject_id, school=school).first()
                if not subject_for_entry:
                    continue
            else:
                subject_for_entry = Subject.objects.filter(id=subject_ids_for_entry[0], school=school).first()
                if not subject_for_entry:
                    continue

            grade_obj = grading.filter(min_score__lte=marks, max_score__gte=marks).first()
            defaults = {
                "marks": marks,
                "grade": grade_obj.grade_letter if grade_obj else "",
                "points": grade_obj.points if grade_obj else 0,
            }

            try:
                StudentMark.objects.update_or_create(
                    student=student,
                    subject=subject_for_entry,
                    term=term,
                    exam=exam,
                    defaults=defaults,
                )
                saved_count += 1
                logger.info("Saved mark: student=%s subject=%s term=%s exam=%s marks=%s", student.id, subject_for_entry.id, term.id, getattr(exam, 'id', None), marks)
            except Exception as e:
                logger.exception("Failed to save mark for student %s: %s", student_id, e)

        messages.success(request, f"{saved_count} marks saved successfully")

        redirect_url = f"{request.path}?class={class_id}&term={term_id}"

        if stream_id:
            redirect_url += f"&stream={stream_id}"
        if exam_id:
            redirect_url += f"&exam={exam_id}"
        if option and option != "all":
            redirect_url += f"&option={option}"
        if entry_mode:
            redirect_url += f"&entry_mode={entry_mode}"
        if entry_mode != "bulk" and subject_id:
            redirect_url += f"&subject={subject_id}"

        return redirect(redirect_url)

    # =========================
    # RENDER
    # =========================
    return render(request, "exams/enter_marks.html", {
        "classes": classes,
        "subjects": subjects,
        "terms": terms,
        "exams": exams,
        "streams": streams,
        "students": students,
        "selected_class": selected_class,
        "selected_option": selected_option,
        "selected_stream": selected_stream,
        "selected_subject": selected_subject,
        "selected_exam": selected_exam,
        "selected_term": selected_term,
        "selected_entry_mode": selected_entry_mode,
        "existing_marks": existing_marks,
        "exam_closed_warning": exam_closed_warning,
        "exam_window_open": exam_window_open,
    })

def report_center(request):
    
    classes = Class.objects.all()
    exams = Exam.objects.all()

    students = []

    class_id = request.GET.get("class_id")

    if class_id:

        students = Student.objects.filter(class_name_id=class_id)

    context = {
        "classes": classes,
        "students": students,
        "exams": exams,
    }

    return render(
        request,
        "exams/report_center.html",
        context
    )

    


from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Spacer,
    Paragraph,
    Image
)

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch

from datetime import date

from schools.models import School, Class, Subject, Term
from students.models import Student
from schools.models import GradingPolicy




# =========================================================
# USE YOUR SAVED GRADING SYSTEM
# =========================================================
def get_grade_and_points(school, mark):

    grading = GradingPolicy.objects.filter(
        school=school,
        min_score__lte=mark,
        max_score__gte=mark
    ).first()

    if grading:
        return grading.short_form or grading.grade_letter, grading.points

    return "-", 0


# =========================================================
# EXPORT PDF
# =========================================================
@login_required
def export_marksheet_pdf(request):

    class_id = request.GET.get("class")
    term_id = request.GET.get("term")
    stream_id = request.GET.get("stream")
    exam_id = request.GET.get("exam")

    school = request.user.school

    class_obj = get_object_or_404(
        Class,
        id=class_id,
        school=school
    )

    term_obj = get_object_or_404(
        Term,
        id=term_id,
        school=school
    )

    exam_obj = None
    if exam_id:
        exam_obj = get_object_or_404(
            Exam,
            id=exam_id,
            school=school
        )

    selected_stream = None
    if stream_id:
        selected_stream = get_object_or_404(Stream, id=stream_id, class_group=class_obj)

    if exam_obj:
        subjects = Subject.objects.filter(
            id__in=exam_obj.exam_subjects.values_list("subject_id", flat=True)
        )
    else:
        subjects = Subject.objects.filter(school=school)

    student_query = Student.objects.filter(
        school=school,
        current_class_id=class_id
    )

    if selected_stream:
        student_query = student_query.filter(stream=selected_stream)

    students = student_query.select_related("stream").order_by("name")

    marks = StudentMark.objects.filter(
        student__in=students,
        term_id=term_id,
    ).select_related(
        "student",
        "subject"
    )
    if exam_obj:
        marks = marks.filter(exam=exam_obj)
    # Debugging: log counts to help diagnose missing marks in export
    try:
        logger.info(
            "Export marksheet request: class_id=%s term_id=%s exam_id=%s stream_id=%s students=%s subjects=%s marks=%s",
            class_id,
            term_id,
            exam_id,
            stream_id,
            students.count() if hasattr(students, 'count') else len(students),
            subjects.count() if hasattr(subjects, 'count') else len(subjects),
            marks.count() if hasattr(marks, 'count') else len(list(marks)),
        )
    except Exception:
        pass
    mark_map = {
        (mark.student_id, mark.subject_id): mark
        for mark in marks
    }

    if selected_stream:
        class_teacher_assignment = ClassTeacherAssignment.objects.filter(
            class_obj=class_obj,
            stream=selected_stream
        ).select_related("teacher").first()
    else:
        class_teacher_assignment = ClassTeacherAssignment.objects.filter(
            class_obj=class_obj
        ).select_related("teacher").first()

    # =====================================================
    # RESPONSE
    # =====================================================
    response = HttpResponse(content_type="application/pdf")

    filename_exam = exam_obj.name if exam_obj else "General"
    response["Content-Disposition"] = (
        f'attachment; filename="{class_obj.name}_Marksheet_{term_obj.name}_{filename_exam}_{date.today().year}.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20,
    )

    styles = getSampleStyleSheet()

    elements = []

    # =====================================================
    # SCHOOL HEADER WITH LOGO ON LEFT
    # =====================================================
    school_info = f"""
    <b>{school.name}</b><br/>
    {school.address if school.address else ''}<br/>
    Phone: {school.phone if school.phone else ''}<br/>
    Email: {school.email if school.email else ''}
    """

    header_paragraph = Paragraph(
        school_info,
        styles["Normal"]
    )

    logo = None

    if school.logo:
        logo = _load_reportlab_image(getattr(school.logo, 'url', school.logo), 0.8 * inch, 0.8 * inch)
    try:
            resolved_logo = None
            try:
                resolved_logo = school.logo.url
            except Exception:
                resolved_logo = getattr(school.logo, 'public_id', None) or getattr(school.logo, 'publicId', None)
                logger = logging.getLogger(__name__)
                logger.debug("Resolved school.logo for school %s: %s", getattr(school, 'id', 'unknown'), resolved_logo)
                if not logo and resolved_logo:
                    logo = _load_reportlab_image(resolved_logo, 0.8 * inch, 0.8 * inch)
    except Exception:
                pass

    if logo:
        header_table = Table(
            [[logo, header_paragraph]],
            colWidths=[0.8 * inch, None]
        )
    else:
        header_table = Table(
            [[header_paragraph]],
            colWidths=[None]
        )

    header_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.black),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 8))

    # =====================================================
    # CLASS INFO BELOW HEADER
    # =====================================================
    class_label = class_obj.name
    if selected_stream:
        class_label = f"{class_obj.name} - {selected_stream.name}"
    else:
        class_label = f"{class_obj.name}"

    class_info = Paragraph(
        f"<b>MARKSHEET:</b> {class_label} | <b>TERM:</b> {term_obj.name} | <b>EXAM:</b> {exam_obj.name if exam_obj else 'General'} | <b>YEAR:</b> {date.today().year}",
        styles["Normal"]
    )
    elements.append(class_info)
    elements.append(Spacer(1, 10))
    elements.append(Spacer(1, 12))

    # =====================================================
    # TABLE HEADER
    # =====================================================
    table_data = []

    header = [
        "ADM NO",
        "NAME",
        "STREAM",
        "GENDER"
    ]

    for subject in subjects:
        header.append(subject.short_name.upper())

    header += [
        "TOTAL",
        "AVG",
        "M.GRADE",
        "T.POINTS",
        "M.POINTS",
        "RANK"
    ]

    table_data.append(header)

    # =====================================================
    # BUILD STUDENT RESULTS
    # =====================================================
    ranking_data = []

    for student in students:

        row = [
            student.admission_number,
            student.name,
            student.stream.name if getattr(student, 'stream', None) else '',
            student.gender[0].upper() if getattr(student, 'gender', None) else ''
        ]

        total_marks = 0
        total_subject_points = 0
        subject_count = 0

        for subject in subjects:
            mark = mark_map.get((student.id, subject.id))

            if mark:
                mark_value = int(mark.marks)
                total_marks += mark_value
                subject_count += 1

                grade, points = get_grade_and_points(
                    school,
                    mark_value
                )

                total_subject_points += points
                row.append(mark_value)
            else:
                row.append("-")

        # =================================================
        # TOTALS
        # =================================================
        average = 0

        if subject_count > 0:
            average = total_marks / subject_count

        final_grade, avg_points = get_grade_and_points(
            school,
            average
        )

        ranking_data.append({
            "student": student,
            "row": row,
            "total": total_marks,
            "average": average,
            "grade": final_grade,
            "points": total_subject_points,
            "subject_count": subject_count,
        })

    ranked_data = assign_competition_ranks(
        [item for item in ranking_data if item["subject_count"] > 0],
        lambda item: (item["points"], item["total"]),
        rank_attr="position",
    )
    unranked_data = [item for item in ranking_data if item["subject_count"] == 0]

    for item in ranked_data:
        row = item["row"]
        row += [
            int(item["total"]),
            round(item["average"], 1),
            item["grade"],
            int(item["points"]),
            round(item["points"] / len(subjects), 2) if len(subjects) > 0 else 0,
            item["position"],
        ]

        table_data.append(row)

    for item in unranked_data:
        row = item["row"]
        row += [
            "-",
            "-",
            "-",
            "-",
            "-",
            "-"
        ]
        table_data.append(row)

    # =====================================================
    # CREATE TABLE
    # =====================================================
    table = Table(table_data, repeatRows=1)

    table.setStyle(TableStyle([
        # GRID
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        # HEADER BOLD
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        # FONTSIZE
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        # CENTER ALIGN
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        # VERTICAL MIDDLE
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # PADDING
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        # STUDENT NAME LEFT ALIGN
        ("ALIGN", (1, 1), (1, -1), "LEFT"),
    ]))

    elements.append(table)

    # =====================================================
    # FOOTER
    # =====================================================
    elements.append(Spacer(1, 20))

    footer = Paragraph(
        f"""
        Generated by Brainet Analytics School System<br/>
        Printed on: {date.today()}
        """,
        styles["Normal"]
    )

    elements.append(footer)

    # =====================================================
    # GRADE ANALYSIS (NEW PAGE)
    # =====================================================
    elements.append(PageBreak())

    analysis_title = Paragraph("<b>GRADE ANALYSIS</b>", styles["Heading2"])
    elements.append(analysis_title)
    elements.append(Spacer(1, 15))

    # Separate by gender
    males = [r for r in ranking_data if getattr(r["student"], "gender", "").lower() in ["m", "male"]]
    females = [r for r in ranking_data if getattr(r["student"], "gender", "").lower() in ["f", "female"]]

    # =====================================================
    # TOP 5 BOYS TABLE
    # =====================================================
    elements.append(Paragraph("<b>TOP 5 BOYS</b>", styles["Heading3"]))
    elements.append(Spacer(1, 8))
    
    top5_boys_data = [["ADM NO", "NAME", "TOTAL", "AVG", "GRADE"]]
    for item in males[:5]:
        s = item["student"]
        top5_boys_data.append([
            s.admission_number,
            s.name,
            int(item["total"]),
            round(item["average"], 1),
            item["grade"]
        ])

    boys_table = Table(top5_boys_data, colWidths=[80, 150, 70, 60, 60])
    boys_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(boys_table)
    elements.append(Spacer(1, 20))

    # =====================================================
    # TOP 5 GIRLS TABLE
    # =====================================================
    elements.append(Paragraph("<b>TOP 5 GIRLS</b>", styles["Heading3"]))
    elements.append(Spacer(1, 8))
    
    top5_girls_data = [["ADM NO", "NAME", "TOTAL", "AVG", "GRADE"]]
    for item in females[:5]:
        s = item["student"]
        top5_girls_data.append([
            s.admission_number,
            s.name,
            int(item["total"]),
            round(item["average"], 1),
            item["grade"]
        ])

    girls_table = Table(top5_girls_data, colWidths=[80, 150, 70, 60, 60])
    girls_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(girls_table)
    elements.append(Spacer(1, 20))

    # =====================================================
    # GRADE DISTRIBUTION TABLE
    # =====================================================
    elements.append(Paragraph("<b>OVERALL GRADE DISTRIBUTION</b>", styles["Heading3"]))
    elements.append(Spacer(1, 8))
    
    grade_counts = {}
    for item in ranking_data:
        g = item["grade"]
        grade_counts[g] = grade_counts.get(g, 0) + 1

    grade_dist_data = [["GRADE", "COUNT"]]
    for k, v in sorted(grade_counts.items(), reverse=True):
        grade_dist_data.append([k, v])

    grade_table = Table(grade_dist_data, colWidths=[100, 80])
    grade_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(grade_table)
    elements.append(Spacer(1, 20))

    # =====================================================
    # BEST PER SUBJECT TABLE
    # =====================================================
    elements.append(Paragraph("<b>BEST PERFORMER PER SUBJECT</b>", styles["Heading3"]))
    elements.append(Spacer(1, 8))
    
    best_data = [["SUBJECT", "BEST STUDENT", "ADM NO", "MARK"]]

    for subject in subjects:
        subj_marks = marks.filter(subject=subject)
        if subj_marks.exists():
            best_mark_obj = subj_marks.order_by('-marks').first()
            best_data.append([
                subject.name,
                best_mark_obj.student.name,
                best_mark_obj.student.admission_number,
                int(best_mark_obj.marks)
            ])
        else:
            best_data.append([subject.name, "-", "-", "-"])

    best_table = Table(best_data, colWidths=[100, 150, 80, 60])
    best_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(best_table)
    elements.append(Spacer(1, 20))

    # =====================================================
    # SUBJECT GRADE DISTRIBUTION TABLE
    # =====================================================
    elements.append(Paragraph("<b>SUBJECT GRADE DISTRIBUTION</b>", styles["Heading3"]))
    elements.append(Spacer(1, 8))
    
    all_grade_letters = set()
    subject_grade_counts = {}

    for subject in subjects:
        subj_marks = marks.filter(subject=subject)
        sg = {}
        for m in subj_marks:
            short_form, _ = get_grade_and_points(school, int(m.marks))
            sg[short_form] = sg.get(short_form, 0) + 1
            all_grade_letters.add(short_form)
        subject_grade_counts[subject.name] = sg

    grade_columns = sorted(all_grade_letters, reverse=True)

    subj_grade_data = [["SUBJECT"] + grade_columns]
    for subj in subjects:
        row = [subj.name]
        sg = subject_grade_counts.get(subj.name, {})
        for g in grade_columns:
            row.append(sg.get(g, 0))
        subj_grade_data.append(row)

    col_widths = [100] + [60] * len(grade_columns)
    subj_table = Table(subj_grade_data, colWidths=col_widths)
    subj_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(subj_table)

    # BUILD PDF
    # =====================================================
    doc.build(elements)

    return response


@login_required
def view_full_marksheet(request):

    # Similar inputs to preview/export but render full marksheet in HTML
    school = request.user.school

    class_id = request.GET.get("class")
    term_id = request.GET.get("term")
    exam_id = request.GET.get("exam")
    stream_id = request.GET.get("stream")

    class_obj = get_object_or_404(Class, id=class_id, school=school)
    term_obj = get_object_or_404(Term, id=term_id, school=school)

    exam_obj = None
    if exam_id:
        exam_obj = Exam.objects.filter(id=exam_id, school=school).first()

    selected_stream = None
    if stream_id:
        selected_stream = get_object_or_404(Stream, id=stream_id, class_group=class_obj)

    if exam_obj:
        subjects = Subject.objects.filter(id__in=exam_obj.exam_subjects.values_list("subject_id", flat=True)).order_by("name")
    else:
        subjects = Subject.objects.filter(school=school).order_by("name")

    student_query = Student.objects.filter(school=school, current_class_id=class_id)
    if selected_stream:
        student_query = student_query.filter(stream=selected_stream)

    students = student_query.select_related("stream").order_by("name")

    marks_qs = StudentMark.objects.filter(student__in=students, term_id=term_id).select_related("subject")
    if exam_obj:
        marks_qs = marks_qs.filter(exam=exam_obj)

    mark_map = {(m.student_id, m.subject_id): m for m in marks_qs}

    report_rows = []

    for student in students:
        subject_scores = []
        total_marks = 0
        total_subjects = 0
        total_points = 0

        for subject in subjects:
            mark = mark_map.get((student.id, subject.id))
            if mark:
                mval = int(mark.marks)
                total_marks += mval
                total_subjects += 1
                if mark.points:
                    total_points += mark.points
            else:
                mval = None
            subject_scores.append(mval)

        average = round(total_marks / total_subjects, 2) if total_subjects > 0 else 0

        grade_obj = GradingPolicy.objects.filter(school=school, min_score__lte=average, max_score__gte=average).first()

        report_rows.append({
            "student": student,
            "stream_name": student.stream.name if student.stream else "",
            "subject_scores": subject_scores,
            "total": total_marks,
            "average": average,
            "grade": grade_obj.short_form if grade_obj else "-",
            "points": total_points,
        })

    report_rows = assign_competition_ranks(
        report_rows,
        lambda row: (row["points"], row["total"]),
        rank_attr="position",
    )

    streams_by_rank = defaultdict(list)
    for row in report_rows:
        streams_by_rank[row["stream_name"]].append(row)

    for stream_rows in streams_by_rank.values():
        assign_competition_ranks(
            stream_rows,
            lambda row: (row["points"], row["total"]),
            rank_attr="stream_rank",
        )

    # analysis: top performers, grade distribution
    ranking_data = report_rows

    males = [r for r in ranking_data if getattr(r["student"], "gender", "").lower() in ["m", "male"]]
    females = [r for r in ranking_data if getattr(r["student"], "gender", "").lower() in ["f", "female"]]

    grade_counts = {}
    for r in ranking_data:
        g = r.get("grade", "-")
        grade_counts[g] = grade_counts.get(g, 0) + 1

    context = {
        "school": school,
        "class_obj": class_obj,
        "term_obj": term_obj,
        "exam_obj": exam_obj,
        "subjects": subjects,
        "report_rows": report_rows,
        "males": males,
        "females": females,
        "grade_counts": grade_counts,
        "display_title": f"{class_obj.name} - Full Marksheet",
        "selected_stream": selected_stream,
    }

    return render(request, "dos/marksheet_full.html", context)


@login_required
def marksheet_center(request):
    
    school = request.user.school

    classes = Class.objects.filter(
        school=school
    )

    terms = Term.objects.filter(
        school=school
    )

    exams = Exam.objects.filter(
        school=school
    )

    subjects = Subject.objects.filter(
        school=school
    )

    selected_class = request.GET.get("class")
    selected_term = request.GET.get("term")
    selected_exam = request.GET.get("exam")
    selected_stream = request.GET.get("stream")

    selected_exam_obj = None
    if selected_exam:
        selected_exam_obj = Exam.objects.filter(id=selected_exam, school=school).first()

    students = []
    streams = []
    selected_stream_obj = None
    display_title = ""

    # =====================================================
    # GET STREAMS FOR SELECTED CLASS
    # =====================================================
    if selected_class:
        streams = Stream.objects.filter(
            class_group_id=selected_class
        ).order_by("name")
        
        if selected_stream:
            selected_stream_obj = get_object_or_404(
                Stream,
                id=selected_stream,
                class_group_id=selected_class
            )

    # =====================================================
    # LOAD STUDENTS + RESULTS
    # =====================================================
    if selected_class and selected_term:

        students = Student.objects.filter(
            school=school,
            current_class_id=selected_class
        ).prefetch_related(
            "studentmark_set"
        )
        
        # Filter by stream if selected
        if selected_stream_obj:
            students = students.filter(stream=selected_stream_obj)
            class_obj = get_object_or_404(Class, id=selected_class)
            display_title = f"{class_obj.name} {selected_stream_obj.name} - Marksheet"
        else:
            class_obj = get_object_or_404(Class, id=selected_class)
            display_title = f"{class_obj.name} - Marksheet"

        # ===============================================
        # PROCESS RESULTS
        # ===============================================
        for student in students:

            total_marks = 0
            total_points = 0
            subject_count = 0

            results = []

            for subject in subjects:

                mark_filter = {
                    "student": student,
                    "subject": subject,
                    "term_id": selected_term,
                }
                if selected_exam_obj is not None:
                    mark_filter["exam"] = selected_exam_obj

                mark_obj = StudentMark.objects.filter(**mark_filter).first()

                if mark_obj:

                    mark = int(mark_obj.marks)

                    total_marks += mark
                    total_points += mark_obj.points
                    subject_count += 1

                    results.append({
                        "subject": subject.name,
                        "marks": mark,
                        "grade": mark_obj.grade,
                        "points": mark_obj.points
                    })

                else:

                    results.append({
                        "subject": subject.name,
                        "marks": "-",
                        "grade": "-",
                        "points": "-"
                    })

            # ===========================================
            # AVERAGE
            # ===========================================
            average = 0

            if subject_count > 0:
                average = round(total_marks / subject_count)

            # ===========================================
            # FINAL GRADE USING YOUR GRADING POLICY
            # ===========================================
            grade_obj = GradingPolicy.objects.filter(
                school=school,
                min_score__lte=average,
                max_score__gte=average
            ).first()

            final_grade = "-"

            if grade_obj:
                final_grade = grade_obj.short_form

            # ===========================================
            # ATTACH TO STUDENT
            # ===========================================
            student.results = results
            student.total_marks = total_marks
            student.average_marks = average
            student.final_grade = final_grade
            student.total_points = total_points
            student.subject_count = subject_count
            student.has_marks = subject_count > 0

        # ===============================================
        # RANKING
        # ===============================================
        ranked_students = assign_competition_ranks(
            [student for student in students if student.has_marks],
            lambda student: student.total_marks,
            rank_attr="rank",
        )

        unranked_students = [student for student in students if not student.has_marks]
        for student in unranked_students:
            student.rank = None

        students = ranked_students + unranked_students

    # =====================================================
    # RENDER PAGE
    # =====================================================
    return render(request, "exams/marksheet_center.html", {

        "classes": classes,
        "terms": terms,
        "exams": exams,
        "subjects": subjects,
        "streams": streams,

        "students": students,

        "selected_class": selected_class,
        "selected_term": selected_term,
        "selected_exam": selected_exam,
        "selected_stream": selected_stream,
        "selected_stream_obj": selected_stream_obj,
        "display_title": display_title,

    })
    
@login_required
def marks_hub(request):

    school = request.user.school

    # =====================================================
    # FILTER DATA
    # =====================================================
    classes = Class.objects.filter(
        school=school
    )

    subjects = Subject.objects.filter(
        school=school
    )

    terms = Term.objects.filter(
        school=school
    )

    exams = Exam.objects.filter(
        school=school
    )

    # =====================================================
    # SELECTED VALUES
    # =====================================================
    selected_class = request.GET.get("class")
    selected_term = request.GET.get("term")
    selected_exam = request.GET.get("exam")
    selected_exam_obj = None
    if selected_exam:
        selected_exam_obj = Exam.objects.filter(id=selected_exam, school=school).first()

    students = []
    subject_headers = []

    # =====================================================
    # LOAD RESULTS
    # =====================================================
    if selected_class and selected_term:

        students = Student.objects.filter(
            school=school,
            current_class_id=selected_class
        )

        subject_headers = Subject.objects.filter(
            school=school
        )

        processed_students = []

        for student in students:

            marks_data = []

            total_marks = 0
            total_points = 0
            subject_count = 0

            # =============================================
            # SUBJECT LOOP
            # =============================================
            for subject in subject_headers:

                mark_filter = {
                    "student": student,
                    "subject": subject,
                    "term_id": selected_term,
                }
                if selected_exam_obj is not None:
                    mark_filter["exam"] = selected_exam_obj

                mark_obj = StudentMark.objects.filter(**mark_filter).first()

                if mark_obj:

                    mark = int(mark_obj.marks)

                    marks_data.append({
                        "subject": subject.name,
                        "marks": mark,
                        "grade": mark_obj.grade,
                        "points": mark_obj.points
                    })

                    total_marks += mark
                    total_points += mark_obj.points
                    subject_count += 1

                else:

                    marks_data.append({
                        "subject": subject.name,
                        "marks": "-",
                        "grade": "-",
                        "points": "-"
                    })

            # =============================================
            # AVERAGE
            # =============================================
            average = 0

            if subject_count > 0:
                average = round(total_marks / subject_count)

            # =============================================
            # FINAL GRADE
            # =============================================
            grade_obj = GradingPolicy.objects.filter(
                school=school,
                min_score__lte=average,
                max_score__gte=average
            ).first()

            final_grade = "-"

            if grade_obj:
                final_grade = grade_obj.grade_letter

            processed_students.append({

                "student": student,

                "marks_data": marks_data,

                "total_marks": total_marks,

                "average": average,

                "grade": final_grade,

                "points": total_points,

            })

        # =================================================
        # RANKING
        # =================================================
        ranked = assign_competition_ranks(
            [p for p in processed_students if p["total_marks"] > 0],
            lambda item: item["total_marks"],
            rank_attr="rank",
        )

        unranked = [p for p in processed_students if p["total_marks"] == 0]

        for item in unranked:
            item["rank"] = None

        students = ranked + unranked

    # =====================================================
    # RENDER
    # =====================================================
    return render(request, "exams/marks_hub.html", {

        "classes": classes,
        "subjects": subjects,
        "terms": terms,
        "exams": exams,

        "students": students,
        "subject_headers": subject_headers,

        "selected_class": selected_class,
        "selected_term": selected_term,
        "selected_exam": selected_exam,
    })

# =====================================================
# IMPORT API VIEWS
# =====================================================
from .views_class_api_patch import class_details_json  
# =========================================================
# MANAGE GRADING POLICIES
# =========================================================

@login_required
def manage_grading(request):

    school = request.user.school

    gradings = GradingPolicy.objects.filter(
        school=school
    ).order_by("-min_score")

    return render(
        request,
        "dos/manage_grading.html",
        {
            "gradings": gradings
        }
    )


# =========================================================
# ADD GRADING POLICY
# =========================================================

@login_required
def add_grading_policy(request):

    school = request.user.school

    if request.method == "POST":

        GradingPolicy.objects.create(
            school=school,
            grade_letter=request.POST.get("grade_letter"),
            short_form=request.POST.get("short_form"),
            min_score=request.POST.get("min_score"),
            max_score=request.POST.get("max_score"),
            points=request.POST.get("points"),
            remarks=request.POST.get("remarks")
        )

        messages.success(
            request,
            "Grading policy added successfully."
        )

    return redirect("manage_grading")


# =========================================================
# EDIT GRADING POLICY
# =========================================================

@login_required
def edit_grading_policy(request, grading_id):

    school = request.user.school

    grading = get_object_or_404(
        GradingPolicy,
        id=grading_id,
        school=school
    )

    if request.method == "POST":

        grading.grade_letter = request.POST.get("grade_letter")
        grading.short_form = request.POST.get("short_form")
        grading.min_score = request.POST.get("min_score")
        grading.max_score = request.POST.get("max_score")
        grading.points = request.POST.get("points")
        grading.remarks = request.POST.get("remarks")

        grading.save()

        messages.success(
            request,
            "Grading policy updated successfully."
        )

        return redirect("manage_grading")

    return render(
        request,
        "dos/edit_grading.html",
        {
            "grading": grading
        }
    )


# =========================================================
# DELETE GRADING POLICY
# =========================================================

@login_required
def delete_grading_policy(request, grading_id):

    school = request.user.school

    grading = get_object_or_404(
        GradingPolicy,
        id=grading_id,
        school=school
    )

    grading.delete()

    messages.success(
        request,
        "Grading policy deleted successfully."
    )

    return redirect("manage_grading")              
                                  
import qrcode, base64
from io import BytesIO


import matplotlib.pyplot as plt

def generate_progress_chart(scores, labels=None, title="Performance Progress", chart_type="line", ymax=100):
    fig, ax = plt.subplots(figsize=(5.8, 2.8), dpi=140)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("#f8fafc")

    if labels:
        x_values = labels
    else:
        x_values = list(range(1, len(scores) + 1))

    if chart_type == "bar":
        colors = ["#2563eb", "#10b981", "#f59e0b", "#ef4444", "#8b5cf6", "#0f766e"]
        bars = ax.bar(x_values, scores, color=colors[:len(scores)] if len(scores) <= len(colors) else None, edgecolor="#1e293b", linewidth=0.8)
        ax.set_ylim(0, ymax)
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2, height + 1, f"{int(height)}", ha="center", va="bottom", fontsize=7, color="#334155")
    else:
        ax.plot(x_values, scores, marker="o", linewidth=2.2, color="#2563eb", markersize=5)
        ax.fill_between(x_values, scores, 0, color="#2563eb", alpha=0.12)
        ax.set_ylim(0, ymax)

    ax.set_title(title, fontsize=9, fontweight="bold", color="#0f172a")
    ax.set_xlabel("Period" if chart_type == "line" else "Subject", fontsize=8, color="#475569")
    ax.set_ylabel("Score", fontsize=8, color="#475569")
    ax.grid(True, linestyle="--", linewidth=0.6, alpha=0.35, color="#94a3b8")
    ax.set_axisbelow(True)
    for spine in ["top", "right"]:
        ax.spines[spine].set_visible(False)
    for spine in ["left", "bottom"]:
        ax.spines[spine].set_color("#cbd5e1")

    plt.tight_layout()
    buffer = BytesIO()
    plt.savefig(buffer, format="PNG", bbox_inches="tight")
    buffer.seek(0)
    return base64.b64encode(buffer.getvalue()).decode()


def get_student_term_performance_history(student):
    """Return a term-by-term performance trend using the average mark for each term."""
    marks_qs = StudentMark.objects.filter(student=student).select_related("term")
    marks_by_term = defaultdict(list)

    for mark in marks_qs:
        if getattr(mark, "term_id", None):
            marks_by_term[mark.term_id].append(float(mark.marks))

    if not marks_by_term:
        return []

    terms = Term.objects.filter(school=student.school, id__in=marks_by_term.keys()).order_by("start_date")
    history = []

    for term in terms:
        term_marks = marks_by_term.get(term.id, [])
        if term_marks:
            average_mark = round(sum(term_marks) / len(term_marks), 1)
            history.append({
                "term_id": term.id,
                "term_name": term.name,
                "start_date": term.start_date,
                "end_date": term.end_date,
                "score": average_mark,
            })

    return history


def get_combined_mark_for_reporting(student, subject, term_obj, combine_requested=False):
    """Return a combined mark for a subject using exam marks from the same term when requested."""
    if not combine_requested:
        return None

    marks_qs = StudentMark.objects.filter(student=student, subject=subject, term=term_obj)
    if not marks_qs.exists():
        return None

    marks = [float(mark.marks) for mark in marks_qs if mark.marks is not None]
    if not marks:
        return None

    return round(sum(marks) / len(marks), 1)

from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
# Alias Image to RLImage to match usage below
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image as RLImage
# Avoid rebinding Final constants elsewhere by aliasing
from reportlab.lib.pagesizes import A4 as RL_A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch as rl_inch

# helper: grade + points + remarks
def get_grade_points_and_remarks(school, mark):
    grading = GradingPolicy.objects.filter(
        school=school,
        min_score__lte=mark,
        max_score__gte=mark
    ).first()
    if grading:
        return grading.short_form or grading.grade_letter, grading.points, grading.remarks
    return "-", 0, ""

# helper: QR image
def generate_qr_image(data):
    from io import BytesIO
    try:
        import qrcode
        qr = qrcode.QRCode(box_size=3, border=2)
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        buffer.seek(0)
        return RLImage(buffer, width=60, height=60)
    except Exception:
        # Fallback: return a small blank image if qrcode is unavailable
        try:
            from PIL import Image as PILImage, ImageDraw
            img = PILImage.new("RGB", (60, 60), color="white")
            draw = ImageDraw.Draw(img)
            draw.rectangle([0, 0, 59, 59], outline="black")
            buffer = BytesIO()
            img.save(buffer, format="PNG")
            buffer.seek(0)
            return RLImage(buffer, width=60, height=60)
        except Exception:
            # Last resort: return None
            return None
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Table, TableStyle,
    Spacer, PageBreak
)

@login_required
def export_class_report(request, class_id, term_id, exam_id):

    school = request.user.school

    class_obj = get_object_or_404(Class, id=class_id, school=school)
    term_obj = get_object_or_404(Term, id=term_id, school=school)
    exam_obj = get_object_or_404(Exam, id=exam_id, school=school)

    stream_id = request.GET.get("stream_id")
    selected_stream = None
    if stream_id:
        selected_stream = get_object_or_404(Stream, id=stream_id, class_group=class_obj)

    # Allow combining multiple exam marks for the same term when requested via query params
    combine_requested = False
    combine_param = request.GET.get("combine") or request.GET.get("combine_requested")
    if combine_param is not None:
        combine_requested = str(combine_param).lower() in ["1", "true", "yes", "on"]

    student_query = Student.objects.filter(
        school=school,
        current_class_id=class_id
    )
    if selected_stream:
        student_query = student_query.filter(stream=selected_stream)

    students = student_query.order_by("name")

    # Use exam-specific subjects when an exam is provided
    if exam_obj:
        subjects = Subject.objects.filter(
            id__in=exam_obj.exam_subjects.values_list("subject_id", flat=True)
        ).order_by("name")
    else:
        subjects = Subject.objects.filter(
            school=school
        ).order_by("name")

    # Restrict report forms to students who have marks for the selected term and, when needed, the chosen exam
    if exam_obj:
        if combine_requested:
            student_ids_with_marks = StudentMark.objects.filter(
                student__in=students,
                term=term_obj,
            ).values_list("student_id", flat=True).distinct()
        else:
            student_ids_with_marks = StudentMark.objects.filter(
                student__in=students,
                term=term_obj,
                exam=exam_obj
            ).values_list("student_id", flat=True).distinct()

        students = students.filter(id__in=list(student_ids_with_marks)).order_by("name")

    # Get class teacher and principal info
    class_teacher_name = "Class Teacher"
    class_teacher_phone = ""
    class_teacher_assignment = None

    if selected_stream:
        class_teacher_assignment = ClassTeacherAssignment.objects.filter(
            class_obj=class_obj,
            stream=selected_stream
        ).select_related("teacher").first()
    else:
        class_teacher_assignment = ClassTeacherAssignment.objects.filter(
            class_obj=class_obj
        ).select_related("teacher").first()

    if class_teacher_assignment and class_teacher_assignment.teacher:
        class_teacher_name = class_teacher_assignment.teacher.name
        class_teacher_phone = class_teacher_assignment.teacher.phone or ""

    principal = school.principals.first()
    principal_name = principal.name if principal else "Principal"
    principal_phone = principal.phone if principal else ""

    # Get previous exam for subject positioning
    previous_exam = Exam.objects.filter(
        school=school,
        term=term_obj
    ).exclude(id=exam_id).order_by('-created_at').first()

    # Use the current calendar year for the report filename
    year = datetime.date.today().year

    # Build term date info for the report header
    term_open_date = term_obj.start_date.strftime("%d %b %Y") if term_obj.start_date else "—"
    term_close_date = term_obj.end_date.strftime("%d %b %Y") if term_obj.end_date else "—"
    next_term = Term.objects.filter(school=school, start_date__gt=term_obj.start_date).order_by("start_date").first()
    next_term_open_date = next_term.start_date.strftime("%d %b %Y") if next_term and next_term.start_date else "—"

    response = HttpResponse(content_type="application/pdf")

    class_label = str(class_obj.name or "Class").strip()
    if selected_stream:
        class_label = f"{class_label} {selected_stream.name}".strip()

    term_label = re.sub(r"\s+", " ", str(term_obj.name or "Term")).strip()
    term_label = re.sub(r"^term\s*", "", term_label, flags=re.IGNORECASE)
    filename = f"{class_label} Report forms term {term_label} {year}.pdf"
    filename = re.sub(r"[\\/:*?\"<>|]+", "-", filename)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportHeader", fontSize=12, leading=16, spaceAfter=6))
    styles.add(ParagraphStyle(name="ReportInfo", fontSize=9, leading=14, spaceAfter=4))
    styles.add(ParagraphStyle(name="ReportSmall", fontSize=8, leading=12))
    elements = []

    # =========================
    # RANKING & POSITIONING
    # =========================
    marks_qs = StudentMark.objects.filter(
        student__in=students,
        term=term_obj
    ).select_related("student", "subject")
    if exam_obj and not combine_requested:
        marks_qs = marks_qs.filter(exam=exam_obj)

    if combine_requested:
        combined_mark_map = {}
        for mark in marks_qs:
            key = (mark.student_id, mark.subject_id)
            combined_mark_map.setdefault(key, []).append(float(mark.marks))

        mark_map = {}
        for key, values in combined_mark_map.items():
            average_mark = round(sum(values) / len(values), 1)
            mark_map[key] = type("CombinedMark", (), {"marks": average_mark})()
    else:
        mark_map = {
            (mark.student_id, mark.subject_id): mark
            for mark in marks_qs
        }

    totals = {
        student.id: sum(
            float(mark_map[(student.id, subject.id)].marks)
            for subject in subjects
            if (student.id, subject.id) in mark_map
        )
        for student in students
    }

    ranked_students = assign_competition_ranks(
        list(students),
        lambda student: totals.get(student.id, 0),
        rank_attr="rank",
    )
    ranks = {student.id: getattr(student, "rank", None) for student in ranked_students}

# ======================
# SUBJECT POSITIONS (current exam only)
# ======================
    subject_positions = {}

    # Initialize subject_positions for all students
    for student in students:
        subject_positions[student.id] = {subject.id: None for subject in subjects}

    # Compute rankings once per subject (not once per student)
    for subject in subjects:
        subject_marks_qs = StudentMark.objects.filter(
            subject=subject,
            term=term_obj
        )
        if exam_obj and not combine_requested:
            subject_marks_qs = subject_marks_qs.filter(exam=exam_obj)

        if combine_requested:
            combined_scores = {}
            for mark in subject_marks_qs:
                combined_scores.setdefault(mark.student_id, []).append(float(mark.marks))

            ranking_rows = []
            for student_id, values in combined_scores.items():
                average_mark = round(sum(values) / len(values), 1)
                ranking_rows.append(type("CombinedMark", (), {"student_id": student_id, "marks": average_mark})())

            ranking = assign_competition_ranks(
                ranking_rows,
                lambda mark: float(mark.marks),
                rank_attr="position",
            )
        else:
            ranking = assign_competition_ranks(
                list(subject_marks_qs.order_by('-marks')),
                lambda mark: float(mark.marks),
                rank_attr="position",
            )

        # assign positions for each student for this subject
        for student in students:
            student_mark = next((mark for mark in ranking if getattr(mark, "student_id", None) == student.id), None)
            subject_positions[student.id][subject.id] = getattr(student_mark, "position", None) if student_mark else None
    # =========================
    for idx, student in enumerate(students):

        # ================= HEADER =================
        logo = None
        if school.logo:
            logo = _load_reportlab_image(getattr(school.logo, 'url', school.logo), 0.8*inch, 0.8*inch)

        stream_display = f" | Stream: {student.stream.name}" if student.stream else ""
        dorm_display = f" | Dorm: {student.dormitory.name}" if student.dormitory else ""

        header_text = f"""
        <b>{school.name}</b><br/>
        {school.address or ""}<br/>
        {school.phone or ""} | {school.email or ""}<br/><br/>
        <b>ACADEMIC REPORT FORM</b><br/>
        Class: {class_obj.name}{stream_display} | Term: {term_obj.name} | Exam: {exam_obj.name}<br/>
        School closed on {term_close_date} | Next term opens on {next_term_open_date}
        """

        if logo:
            header = Table(
                [[logo, Paragraph(header_text, styles["Normal"])]],
                colWidths=[70, 440]
            )
        else:
            header = Table(
                [[Paragraph(header_text, styles["Normal"])]],
                colWidths=[None]
            )

        header.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LINEBELOW", (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(header)
        elements.append(Spacer(1, 8))

        # ================= STUDENT INFO =================
        overall_pos = format_position_display(ranks.get(student.id))
        overall_pos_display = f"{overall_pos}/{len(students)}" if overall_pos else f"{len(students)}"
        info = f"""
        <b>Student Name:</b> {student.name} &nbsp;&nbsp;
        <b>Adm No:</b> {student.admission_number} &nbsp;&nbsp;
        <b>Position:</b> {overall_pos_display}{dorm_display}
        """

        elements.append(Paragraph(info, styles["ReportInfo"]))
        elements.append(Spacer(1, 8))

        # ================= MARKS TABLE WITH TEACHER & POSITIONING =================
        table_data = [[
            Paragraph("Subject", styles["ReportSmall"]),
            Paragraph("Marks", styles["ReportSmall"]),
            Paragraph("Grade", styles["ReportSmall"]),
            Paragraph("Points", styles["ReportSmall"]),
            Paragraph("Teacher", styles["ReportSmall"]),
            Paragraph("Pos", styles["ReportSmall"]),
            Paragraph("Comments", styles["ReportSmall"]),
        ]]

        total_marks = 0
        total_points = 0

        for subject in subjects:

            mark_filter = {
                "student": student,
                "subject": subject,
                "term": term_obj,
            }
            if exam_obj and not combine_requested:
                mark_filter["exam"] = exam_obj

            mark = StudentMark.objects.filter(**mark_filter).first()
            if combine_requested and not mark:
                combined_mark = get_combined_mark_for_reporting(student, subject, term_obj, combine_requested=True)
                if combined_mark is not None:
                    mark = type("CombinedMark", (), {"marks": combined_mark, "grade": "", "points": 0})()

            # Get subject teacher
            teacher_name = "—"
            subject_teacher = TeacherSubjectAssignment.objects.filter(
                class_obj=class_obj,
                subject=subject,
                stream=student.stream if student.stream else None
            ).select_related("teacher").first()
            
            if subject_teacher:
                teacher_name = subject_teacher.teacher.name[:15]  # Abbreviate for space

            # Subject positioning indicator
            subject_position = ""
            student_key = getattr(student, "id", None) or getattr(student, "pk", None)
            # subject may not expose an `id` attribute; prefer `id` or `pk` safely
            subject_key = getattr(subject, "id", None) or getattr(subject, "pk", None)
            if student_key and subject_positions.get(student_key):
                subject_position = subject_positions[student_key].get(subject_key)
            subject_position_display = format_position_display(subject_position)
            if mark:
                m = int(round(mark.marks))
                grade, points, remarks = get_grade_points_and_remarks(school, m)

                total_marks += m
                total_points += points

                table_data.append([
                    Paragraph(subject.name, styles["ReportSmall"]),
                    Paragraph(str(m), styles["ReportSmall"]),
                    Paragraph(str(grade), styles["ReportSmall"]),
                    Paragraph(str(points), styles["ReportSmall"]),
                    Paragraph(teacher_name, styles["ReportSmall"]),
                    Paragraph(str(subject_position_display), styles["ReportSmall"]),
                    Paragraph(remarks, styles["ReportSmall"])
                ])
            else:
                table_data.append([
                    Paragraph(subject.name, styles["ReportSmall"]),
                    Paragraph("-", styles["ReportSmall"]),
                    Paragraph("-", styles["ReportSmall"]),
                    Paragraph("-", styles["ReportSmall"]),
                    Paragraph(teacher_name, styles["ReportSmall"]),
                    Paragraph(str(subject_position_display), styles["ReportSmall"]),
                    Paragraph("-", styles["ReportSmall"])
                ])

        # ================= TABLE STYLE =================
        table = Table(
            table_data,
            colWidths=[130, 45, 40, 40, 80, 30, 90]
        )

        table.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 1), (4, -1), "CENTER"),
            ("ALIGN", (5, 1), (5, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 10))

        # ================= TOTALS ANALYSIS =================
        avg = total_marks / len(subjects) if subjects else 0
        avg_display = int(round(avg))
        final_grade, _, final_remarks = get_grade_points_and_remarks(school, avg_display)

        totals_text = (
            f"<b>Total Marks:</b> {int(total_marks)} | "
            f"<b>Average:</b> {avg_display} | "
            f"<b>Total Points:</b> {int(total_points)} | "
            f"<b>Grade:</b> {final_grade}"
        )

        elements.append(Paragraph(totals_text, styles["ReportHeader"]))
        elements.append(Spacer(1, 8))

        # ================= SIGNATURES & COMMENTS SECTION =================
        teacher_block = []
        teacher_block.append(Paragraph(f"<b>Class Teacher: {class_teacher_name}</b>", styles["ReportInfo"]))
        if class_teacher_phone:
            teacher_block.append(Paragraph(f"Phone: {class_teacher_phone}", styles["ReportSmall"]))
            remarks = final_remarks if final_remarks else "—"
            teacher_block.append(Paragraph(f"Comments: {remarks}", styles["ReportSmall"]))
        teacher_block.append(Paragraph("Signature: ________________________", styles["ReportSmall"]))
        teacher_block.append(Spacer(1, 4))
        
        teacher_block.append(Paragraph(f"<b>Principal: {principal_name}</b>", styles["ReportInfo"]))
        if principal_phone:
            teacher_block.append(Paragraph(f"Phone: {principal_phone}", styles["ReportSmall"]))
            remarks = final_remarks if final_remarks else "—"
            teacher_block.append(Paragraph(f"Comments: {remarks}", styles["ReportSmall"]))
        teacher_block.append(Paragraph("Signature: ________________________", styles["ReportSmall"]))
        teacher_block.append(Spacer(1, 6))

        # QR code for student login page
        try:
            login_url = request.build_absolute_uri(reverse("student_login"))
            qr_img = generate_qr_image(login_url)
        except Exception:
            qr_img = None

        # Performance charts for this student
        term_chart_img = None
        subject_chart_img = None
        try:
            term_scores = []
            term_labels = []
            for entry in performance_history:
                if entry.get("term_name"):
                    term_labels.append(entry["term_name"])
                    term_scores.append(float(entry["score"]))

            if term_scores:
                term_b64 = generate_progress_chart(term_scores, labels=term_labels, title=f"{student.name} Performance Over Time", chart_type="line")
                import base64
                from io import BytesIO
                term_bytes = base64.b64decode(term_b64)
                term_buf = BytesIO(term_bytes)
                term_chart_img = RLImage(term_buf, width=240, height=120)
        except Exception:
            term_chart_img = None

        try:
            scores = []
            labels = []
            for subject in subjects:
                m_filter = {"student": student, "subject": subject, "term": term_obj}
                if exam_obj:
                    m_filter["exam"] = exam_obj
                m = StudentMark.objects.filter(**m_filter).first()
                scores.append(int(round(m.marks)) if m else 0)
                labels.append(subject.short_name or subject.name[:8])

            chart_title = f"{student.name} Subject Performance"
            if exam_obj:
                chart_title = f"{student.name} Subject Performance - {exam_obj.name}"
            else:
                chart_title = f"{student.name} Subject Performance - {term_obj.name}"

            b64 = generate_progress_chart(scores, labels, title=chart_title, chart_type="bar")
            import base64
            from io import BytesIO
            img_bytes = base64.b64decode(b64)
            buf = BytesIO(img_bytes)
            subject_chart_img = RLImage(buf, width=240, height=120)
        except Exception:
            subject_chart_img = None

        # Build side-by-side layout
        left_flowables = teacher_block

        right_flowables = []
        if qr_img:
            right_flowables.append(qr_img)
            login_email = (
                student.user.email
                if hasattr(student, 'user') and getattr(student.user, 'email', None)
                else student_login_email(student.admission_number, school)
            )
            login_text = (
                f"Scan the QR code to login.<br/>"
                f"<b>Username:</b> {login_email}<br/>"
                f"<b>Password:</b> {student.admission_number}"
            )
            right_flowables.append(Spacer(1, 4))
            right_flowables.append(Paragraph(login_text, styles["ReportSmall"]))

        if term_chart_img:
            right_flowables.append(Spacer(1, 6))
            right_flowables.append(Paragraph("Performance Over Time", styles["ReportSmall"]))
            right_flowables.append(term_chart_img)

        if subject_chart_img:
            right_flowables.append(Spacer(1, 6))
            right_flowables.append(Paragraph("Subject Performance", styles["ReportSmall"]))
            right_flowables.append(subject_chart_img)

        if not term_chart_img and not subject_chart_img:
            right_flowables.append(Paragraph("Charts unavailable", styles["ReportSmall"]))

        columns_table = Table([[left_flowables, right_flowables]], colWidths=[220, 300])
        columns_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.0, colors.white),
        ]))

        elements.append(columns_table)
        elements.append(Spacer(1, 8))

        # ================= FOOTER =================
        elements.append(Paragraph(
            "Powered by Brainet Analytics | Report Forms",
            styles["Normal"]
        ))

        # PAGE BREAK
        if idx < len(students) - 1:
            elements.append(PageBreak())

    doc.build(elements)

    return response

@login_required
def clear_notifications(request):
    qs = Notification.objects.filter(recipient=request.user)
    deleted_info = qs.delete()
    deleted_count = deleted_info[0] if isinstance(deleted_info, (list, tuple)) else 0

    try:
        from django.contrib import messages
        messages.success(request, f"Cleared {deleted_count} notifications.")
    except Exception:
        pass

    # If this was an AJAX/API call, return JSON; otherwise redirect back to dashboard
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"deleted": deleted_count})

    return redirect("dos_dashboard")  # adjust to your dashboard view name

def add_term(request):
    if request.method == "POST":
        name = request.POST.get("name")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        school = request.user.school

        Term.objects.create(
            name=name,
            start_date=start_date,
            end_date=end_date,
            school=school
        )

        messages.success(request, "Term added successfully ✅")
        return redirect("manage_terms")

    return render(request, "schools/add_term.html")


@login_required
def edit_term(request, term_id):
    school = request.user.school
    term = get_object_or_404(Term, id=term_id, school=school)

    if request.method == "POST":
        term.name = request.POST.get("name")
        term.start_date = request.POST.get("start_date")
        term.end_date = request.POST.get("end_date")
        term.save()

        messages.success(request, "Term updated successfully ✏️")
        return redirect("manage_terms")

    return render(request, "schools/edit_term.html", {"term": term})


@login_required
def delete_term(request, term_id):
    school = request.user.school
    term = get_object_or_404(Term, id=term_id, school=school)

    if request.method == "POST":
        term.delete()
        messages.success(request, "Term deleted successfully 🗑")
        return redirect("manage_terms")

    return render(request, "schools/delete_term.html", {"term": term})


# =========================================================
# STUDENT PROMOTION
# =========================================================

@login_required
def promotion_center(request):
    """Main promotion center for managing student promotions"""
    school = request.user.school
    
    classes = Class.objects.filter(school=school)
    students = Student.objects.filter(school=school, status='active')
    
    context = {
        'classes': classes,
        'students': students,
        'total_students': students.count(),
    }
    return render(request, "schools/promotion_center.html", context)


@login_required
def promote_class_view(request, class_id):
    """Promote all students in a class to the next level"""
    from .promotion_service import PromotionService
    
    school = request.user.school
    class_obj = get_object_or_404(Class, id=class_id, school=school)
    
    if request.method == "POST":
        stats = PromotionService.promote_class(class_obj, promoted_by=request.user)
        
        messages.success(
            request,
            f"Promotion complete! Promoted: {stats['promoted']}, "
            f"Graduated: {stats['graduated']}, Failed: {stats['failed']}"
        )
        return redirect("promotion_center")
    
    students = Student.objects.filter(current_class=class_obj, status='active')
    
    context = {
        'class': class_obj,
        'students': students,
    }
    return render(request, "schools/promote_class_confirm.html", context)


@login_required
def promote_student_view(request, student_id):
    """Promote or handle individual student"""
    from .promotion_service import PromotionService
    
    school = request.user.school
    student = get_object_or_404(Student, id=student_id, school=school)
    
    if request.method == "POST":
        action = request.POST.get("action", "promote")
        remarks = request.POST.get("remarks", "")
        
        if action == "promote":
            next_class, next_stream = PromotionService.get_next_class(
                student.current_class, 
                student.stream
            )
            promotion = PromotionService.promote_student(
                student,
                next_class,
                next_stream,
                promoted_by=request.user,
                remarks=remarks
            )
            messages.success(
                request,
                f"{student.name} promoted to {next_class.name if next_class else 'Graduated'}"
            )
        
        elif action == "repeat":
            PromotionService.repeat_student(
                student,
                repeated_by=request.user,
                remarks=remarks
            )
            messages.success(request, f"{student.name} will repeat the year")
        
        elif action == "drop":
            PromotionService.drop_student(
                student,
                dropped_by=request.user,
                remarks=remarks
            )
            messages.success(request, f"{student.name} has been dropped from school")
        
        return redirect("promotion_center")
    
    next_class, next_stream = PromotionService.get_next_class(
        student.current_class,
        student.stream
    )
    
    context = {
        'student': student,
        'next_class': next_class,
        'next_stream': next_stream,
    }
    return render(request, "schools/promote_student.html", context)


@login_required
def promote_school_view(request):
    """Promote all students in the school"""
    from .promotion_service import PromotionService
    
    school = request.user.school
    
    if request.method == "POST":
        stats = PromotionService.promote_school(school, promoted_by=request.user)
        
        messages.success(
            request,
            f"School-wide promotion complete! Total: {stats['total_students']}, "
            f"Promoted: {stats['promoted']}, Graduated: {stats['graduated']}, "
            f"Failed: {stats['failed']}"
        )
        return redirect("promotion_center")
    
    classes = Class.objects.filter(school=school)
    students_count = Student.objects.filter(school=school, status='active').count()
    
    context = {
        'classes': classes,
        'students_count': students_count,
    }
    return render(request, "schools/promote_school_confirm.html", context)


@login_required
def promotion_history(request):
    """View promotion history"""
    school = request.user.school
    promotions = StudentPromotion.objects.filter(school=school).select_related(
        'student', 'from_class', 'to_class', 'promoted_by'
    )
    
    # Filtering
    student_name = request.GET.get("student")
    status_filter = request.GET.get("status")
    
    if student_name:
        promotions = promotions.filter(student__name__icontains=student_name)
    
    if status_filter:
        promotions = promotions.filter(status=status_filter)
    
    context = {
        'promotions': promotions,
        'student_name': student_name,
        'status_filter': status_filter,
        'status_choices': StudentPromotion.PROMOTION_STATUS,
    }
    return render(request, "schools/promotion_history.html", context)


@login_required
def undo_promotion(request, promotion_id):
    """Undo a previous promotion record and restore the student's previous class/stream."""
    promotion = get_object_or_404(StudentPromotion, id=promotion_id, school=request.user.school)
    student = promotion.student

    if request.method != "POST":
        messages.error(request, "Invalid undo request.")
        return redirect("promotion_history")

    if promotion.status not in ["promoted", "graduated"]:
        messages.error(request, "Only promoted or graduated records can be undone.")
        return redirect("promotion_history")

    if not promotion.from_class:
        messages.error(request, "This promotion cannot be undone because the previous class is not available.")
        return redirect("promotion_history")

    # Ensure the student is still in the promotion target state before undoing
    if promotion.status == "promoted":
        if student.current_class != promotion.to_class or student.stream != promotion.to_stream:
            messages.error(request, "Student is no longer in the promoted class/stream. Undo not possible.")
            return redirect("promotion_history")

    if promotion.status == "graduated":
        if student.current_class is not None:
            messages.error(request, "Student is no longer graduated. Undo not possible.")
            return redirect("promotion_history")

    student.current_class = promotion.from_class
    student.stream = promotion.from_stream
    student.status = 'active'
    student.save()

    promotion.remarks = (promotion.remarks or "") + " [Undo performed]"
    promotion.save()

    messages.success(request, f"Promotion for {student.name} has been undone.")
    return redirect("promotion_history")


# =========================================================
# MANUAL USER VERIFICATION - Admin Helper
# =========================================================
@login_required
@superuser_required
def manual_verify_user(request, user_id):
    """
    Manually verify a user's email account.
    Helps bypass failed verification attempts.
    Only accessible to superusers.
    """
    user = get_object_or_404(User, id=user_id)
    
    if request.method == "POST":
        try:
            user.mark_email_verified()
            messages.success(
                request,
                f"✅ User {user.email} has been manually verified!"
            )
            return redirect("superuser_dashboard")
        except Exception as e:
            messages.error(
                request,
                f"❌ Error verifying user: {str(e)}"
            )
            return redirect("superuser_dashboard")
    
    context = {
        'user': user,
        'is_verified': user.email_verified,
    }
    return render(request, "schools/manual_verify_user.html", context)


# =========================================================
# GET ONLINE CLASS MEETING LINK - With Authorization
# =========================================================
@login_required
def get_online_class_meeting_link(request, online_class_id):
    """
    Fetch the meeting link for an online class.
    ONLY returns the link if student is in the target class/stream.
    
    This prevents unauthorized students from accessing meeting links.
    """
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        return JsonResponse(
            {"error": "Student profile not found"},
            status=403
        )
    
    online_class = get_object_or_404(OnlineClass, id=online_class_id)
    
    # =========================================
    # VERIFY STUDENT IS IN TARGET CLASS
    # =========================================
    if student.current_class_id != online_class.class_obj_id:
        return JsonResponse(
            {"error": "You are not enrolled in this class"},
            status=403
        )
    
    # =========================================
    # VERIFY STUDENT IS IN TARGET STREAM (if stream-specific)
    # =========================================
    if online_class.stream:
        if not student.stream or student.stream_id != online_class.stream_id:
            return JsonResponse(
                {"error": "You are not in the stream for this class"},
                status=403
            )
    
    # =========================================
    # RETURN MEETING LINK (Authorized)
    # =========================================
    from django.http import JsonResponse
    return JsonResponse({
        "success": True,
        "meeting_link": online_class.meeting_link,
        "topic": online_class.topic,
        "teacher": online_class.teacher.name,
    })


# =========================================================
# STUDENT JOIN ONLINE CLASS
# =========================================================
@login_required
def student_join_online_class(request, online_class_id):
    """
    Allow a student to join an online class.
    Creates/updates OnlineClassParticipant record and redirects to meeting link.
    
    TARGETING LOGIC:
    - Student MUST be in the online_class.class_obj
    - IF online_class.stream is set: student MUST be in that stream
    - IF online_class.stream is NULL: all students in class can join
    """
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, "❌ Student profile not found.")
        return redirect("student_dashboard")
    
    online_class = get_object_or_404(OnlineClass, id=online_class_id)
    
    # =========================================
    # VERIFY STUDENT IS IN TARGET CLASS
    # =========================================
    if student.current_class_id != online_class.class_obj_id:
        messages.error(
            request,
            f"❌ You are not in {online_class.class_obj.name}. This class is only for that class."
        )
        return redirect("student_dashboard")
    
    # =========================================
    # VERIFY STUDENT IS IN TARGET STREAM (if stream-specific)
    # =========================================
    if online_class.stream:
        if not student.stream or student.stream_id != online_class.stream_id:
            messages.error(
                request,
                f"❌ You are not in {online_class.stream.name}. This class is only for that stream."
            )
            return redirect("student_dashboard")
    
    # =========================================
    # CREATE/UPDATE PARTICIPANT RECORD
    # =========================================
    try:
        participant, created = OnlineClassParticipant.objects.get_or_create(
            online_class=online_class,
            student=student
        )
        
        # Mark as joined
        participant.status = "joined"
        participant.save()
        
        action = "successfully joined" if created else "already joined"
        messages.success(
            request,
            f"✅ You have {action} '{online_class.topic}'! Redirecting to meeting..."
        )
        
        # =========================================
        # REDIRECT TO MEETING LINK (Authorized Access Only)
        # =========================================
        if online_class.meeting_link:
            return redirect(online_class.meeting_link)
        else:
            messages.warning(
                request,
                "⚠️ No meeting link available yet. Please contact your teacher."
            )
            return redirect("student_dashboard")
            
    except Exception as e:
        messages.error(
            request,
            f"❌ Error joining class: {str(e)}"
        )
        return redirect("student_dashboard")


# =========================================================
# STUDENT ONLINE CLASS ACTION (For other interactions)
# =========================================================
@login_required
def student_online_class_action(request, online_class_id):
    """
    Handle student actions on online classes (view details, record attempt, etc)
    
    TARGETING LOGIC:
    - Validates student is in the target class
    - Validates student is in target stream (if stream-specific)
    - Does NOT expose meeting link directly
    """
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect("student_dashboard")
    
    online_class = get_object_or_404(OnlineClass, id=online_class_id)
    
    # =========================================
    # VERIFY STUDENT CAN ACCESS THIS CLASS
    # =========================================
    if student.current_class_id != online_class.class_obj_id:
        messages.error(request, "❌ You don't have access to this class.")
        return redirect("student_dashboard")
    
    if online_class.stream and (not student.stream or student.stream_id != online_class.stream_id):
        messages.error(request, "❌ You don't have access to this stream's class.")
        return redirect("student_dashboard")
    
    # Get or create participant record
    participant, _ = OnlineClassParticipant.objects.get_or_create(
        online_class=online_class,
        student=student
    )
    
    action = request.GET.get("action", "view")
    
    if action == "view":
        # Show details WITHOUT the meeting link
        context = {
            'online_class': online_class,
            'participant': participant,
            'meeting_link': None,  # Don't expose link here
        }
        return render(request, "students/online_class_detail.html", context)
    
    elif action == "failed":
        # Mark as failed attempt
        participant.status = "failed"
        participant.save()
        messages.warning(request, "Failed join attempt recorded. Please try again or contact your teacher.")
        return redirect("student_dashboard")
    
    else:
        return redirect("student_dashboard")


# =========================================================
# MANUAL JOIN ONLINE CLASS - Admin Helper
# =========================================================
@login_required
@superuser_required
def manual_join_online_class(request, online_class_id, student_id):
    """
    Manually add a student to an online class and mark them as joined.
    Helps bypass failed join attempts.
    Only accessible to superusers.
    """
    online_class = get_object_or_404(OnlineClass, id=online_class_id)
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == "POST":
        try:
            participant, created = OnlineClassParticipant.objects.get_or_create(
                online_class=online_class,
                student=student
            )
            
            # Mark as joined
            participant.status = "joined"
            participant.save()
            
            action = "joined" if created else "updated to joined"
            messages.success(
                request,
                f"✅ {student.name} has been manually {action} the online class '{online_class.topic}'!"
            )
            return redirect("superuser_dashboard")
        except Exception as e:
            messages.error(
                request,
                f"❌ Error adding student to class: {str(e)}"
            )
            return redirect("superuser_dashboard")
    
    context = {
        'online_class': online_class,
        'student': student,
    }
    return render(request, "schools/manual_join_online_class.html", context)